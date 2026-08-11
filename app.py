from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for, Response
from jinja2 import ChoiceLoader, FileSystemLoader, DictLoader
from markupsafe import escape
from pathlib import Path
from datetime import datetime
from PIL import Image, ImageOps, ImageEnhance
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from http.server import ThreadingHTTPServer
from urllib.parse import urlsplit
import subprocess, tempfile, csv, io, json, re, uuid, statistics, shutil, os, traceback, unicodedata, threading, http.client, time, hmac, hashlib

import orders_app
from scanner_templates import SCANNER_TEMPLATES, SCANNER_CSS

ROOT = Path(__file__).resolve().parent
TEMPLATES_DIR = ROOT / "templates"
STATIC_DIR = ROOT / "static"
app = Flask(__name__, template_folder=str(TEMPLATES_DIR), static_folder=str(STATIC_DIR))
# Respaldo embebido: si Render no encuentra/copió templates/, Jinja usa estas plantillas incluidas en Python.
app.jinja_loader = ChoiceLoader([
    FileSystemLoader(str(TEMPLATES_DIR)),
    DictLoader(SCANNER_TEMPLATES),
])
app.secret_key = os.environ.get("SECRET_KEY", "cuponera-v19-dual")

# Configuración compartida del portal TALMA.
CONFIG = {"secret_key": app.secret_key}

def esc(value):
    """Escapa valores antes de insertarlos en HTML generado por el portal TALMA."""
    return str(escape("" if value is None else value))

def page(title: str, body: str) -> str:
    """Página HTML base usada por el portal privado TALMA."""
    return f"""<!doctype html>
<html lang=\"es\">
<head>
<meta charset=\"utf-8\">
<meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">
<meta name=\"robots\" content=\"noindex,nofollow\">
<title>{esc(title)} — Urpicha Restaurante</title>
<style>{SCANNER_CSS}
.wrap{{max-width:1200px;margin:30px auto;padding:20px}}
.narrow{{max-width:560px}}
.card{{background:#fff;padding:24px;border:1px solid #d8dfe6;border-radius:15px;box-shadow:0 5px 18px rgba(0,0,0,.05);margin-bottom:20px}}
.actions{{display:flex;gap:10px;align-items:center;flex-wrap:wrap}}
.btn{{display:inline-block;border:0;border-radius:9px;padding:10px 14px;background:#176b43;color:#fff;text-decoration:none;font-weight:700}}
.btn.secondary{{background:#e7edf1;color:#17212b}}
.muted{{color:#5f6b76}}
.notice{{padding:12px 14px;border-radius:8px;background:#e9f6ef;margin:12px 0}}
.notice.error{{background:#fdecea;color:#8a1c13;border-left:5px solid #b42318}}
.grid{{display:grid;gap:16px}}
.grid3{{grid-template-columns:repeat(3,minmax(0,1fr))}}
.stat{{background:#fff;border:1px solid #d8dfe6;border-radius:14px;padding:18px;font-weight:700}}
.stat b{{display:block;font-size:28px;margin-top:6px}}
.table-wrap{{overflow:auto}}
@media(max-width:800px){{.grid3{{grid-template-columns:1fr}}}}
</style>
</head>
<body>{body}</body>
</html>"""
DATA_ROOT = Path(os.environ.get("DATA_DIR", str(ROOT / "datos"))).expanduser().resolve()
SCANNER_ROOT = DATA_ROOT / "scanners"
UPLOADS = SCANNER_ROOT / "uploads"
OUTPUTS = SCANNER_ROOT / "outputs"

UPLOADS.mkdir(parents=True, exist_ok=True)
OUTPUTS.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Configuración por sistema
# ---------------------------------------------------------------------------
SISTEMAS = {
    "talma": {
        "titulo": "TALMA",
        "subtitulo": "Carga la captura del Excel de pedidos TALMA. Antes de leerla podrás ajustar las columnas.",
        "ayuda": "Sistema TALMA: código, nombre, área (RAMPA, PAX, CARGA, OMA), entrada, segundo y observación.",
        "data_file": SCANNER_ROOT / "datos_talma.json",
        "cols": ["registro", "fecha", "sede", "codigo", "nombre", "area", "entrada", "segundo", "observacion"],
        "areas": ["RAMPA", "PAX", "CARGA", "OMA"],
        "label_entrada": "ENTRADA",
        "label_segundo": "SEGUNDO",
        "label_observacion": "OBSERVACIÓN",
        "default_entry": 0.455,
        "default_second": 0.545,
        "default_observation": 0.685,
        "excel_headers": ["Registro", "Fecha", "Sede", "Código", "Nombre", "N° Almuerzo", "Área", "Entrada", "Segundo", "Observación"],
        "excel_widths": [24, 15, 12, 14, 30, 14, 12, 28, 34, 50],
        "pdf_prefix": "talma",
        "header_label": None,  # usa área
    },
    "policia": {
        "titulo": "POLICÍA",
        "subtitulo": "Carga la captura de la planilla PNP. Ajusta las barras de Entrada, Plato de fondo y Sugerencias.",
        "ayuda": "Sistema POLICÍA: N°, nombre y apellido, entrada, plato de fondo y sugerencias. Sin columna de área.",
        "data_file": SCANNER_ROOT / "datos_policia.json",
        "cols": ["numero", "nombre", "entrada", "segundo", "observacion"],
        "areas": [],
        "label_entrada": "ENTRADA",
        "label_segundo": "PLATO DE FONDO",
        "label_observacion": "SUGERENCIAS",
        # Formato PNP de la captura de Excel: C empieza ~33%, D ~56.6%, E ~79.8%.
        "default_entry": 0.330,
        "default_second": 0.566,
        "default_observation": 0.798,
        "excel_headers": ["N°", "Nombre y apellido", "N° Almuerzo", "Entrada", "Plato de fondo", "Sugerencias"],
        "excel_widths": [8, 42, 14, 28, 34, 28],
        "pdf_prefix": "policia",
        "header_label": "POLICÍA",
    },
}

# Compatibilidad: migrar datos antiguos si existen
_legacy = SCANNER_ROOT / "datos_actuales.json"
_talma_data = SISTEMAS["talma"]["data_file"]
if _legacy.exists() and not _talma_data.exists():
    try:
        shutil.copy(_legacy, _talma_data)
    except OSError:
        pass


def cfg(sistema: str) -> dict:
    key = (sistema or "").strip().lower()
    if key not in SISTEMAS:
        raise RuntimeError(f"Sistema desconocido: {sistema}")
    return SISTEMAS[key]


def repair_mojibake(value: str) -> str:
    """Repara texto UTF-8 interpretado por error como Windows-1252/Latin-1."""
    value = value or ""
    markers = ("Ã", "Â", "â€", "â€™", "â€œ", "â€\u009d", "ðŸ", "�")
    for _ in range(3):
        if not any(marker in value for marker in markers):
            break
        repaired = None
        for encoding in ("cp1252", "latin-1"):
            try:
                candidate = value.encode(encoding).decode("utf-8")
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
            if candidate != value:
                repaired = candidate
                break
        if repaired is None:
            break
        value = repaired
    return value


def remove_vowel_accents(value: str) -> str:
    """Quita tildes y diéresis, pero conserva correctamente la letra ñ."""
    protected = value.replace("ñ", "__ENYE_LOWER__").replace("Ñ", "__ENYE_UPPER__")
    decomposed = unicodedata.normalize("NFD", protected)
    without_marks = "".join(char for char in decomposed if unicodedata.category(char) != "Mn")
    return without_marks.replace("__ENYE_LOWER__", "ñ").replace("__ENYE_UPPER__", "Ñ")


def clean_text(text: str, preserve_accents: bool = False) -> str:
    """Limpia OCR/codificación. Por compatibilidad TALMA quita tildes salvo que se pidan."""
    value = str(text or "")
    for stray in ("™", "®", "©", "�"):
        value = value.replace(stray, "")
    value = repair_mojibake(value)
    value = unicodedata.normalize("NFC", value)

    replacements = {
        "\u00a0": " ",
        "\u200b": "",
        "\ufeff": "",
        "|": " ",
        "™": "",
        "®": "",
        "©": "",
        "�": "",
        "â€”": "-",
        "â€“": "-",
        "â€™": "'",
        "â€œ": '"',
        "â€\u009d": '"',
        "Ã": "",
        "Â": "",
        "â": "",
    }
    for bad, good in replacements.items():
        value = value.replace(bad, good)

    allowed_punctuation = set(" .,:;!?¿¡'\"()/-+&°#%")
    chars = []
    for char in value:
        category = unicodedata.category(char)
        if char.isspace() or char in allowed_punctuation or category[0] in ("L", "N"):
            chars.append(char)
    value = "".join(chars)
    value = re.sub(r"\s+", " ", value).strip()
    if not preserve_accents:
        value = remove_vowel_accents(value)
    return value


def clean_text_unicode(text: str) -> str:
    """Limpia el texto conservando tildes, diéresis, ñ y signos españoles válidos."""
    return clean_text(text, preserve_accents=True)


def pdf_text(text: str, preserve_accents: bool = False) -> str:
    """Texto seguro para las fuentes PDF incorporadas, conservando acentos cuando corresponde."""
    value = clean_text(text, preserve_accents=preserve_accents)
    return value.encode("cp1252", "ignore").decode("cp1252")


def find_tesseract():
    candidates = [
        shutil.which("tesseract"),
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Programs\Tesseract-OCR\tesseract.exe"),
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)
    return None


def available_languages(exe):
    try:
        result = subprocess.run(
            [exe, "--list-langs"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=15,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
        )
        return [line.strip() for line in result.stdout.splitlines()[1:] if line.strip()]
    except Exception:
        return []


def run_tesseract_tsv(image):
    exe = find_tesseract()
    if not exe:
        raise RuntimeError("Tesseract OCR no está instalado. Ejecuta instalar_ocr.bat.")

    languages = available_languages(exe)
    language = "spa" if "spa" in languages else ("eng" if "eng" in languages else None)

    image = ImageOps.grayscale(image)
    if image.width < 1800:
        ratio = 1800 / image.width
        image = image.resize((1800, int(image.height * ratio)))
    image = ImageEnhance.Contrast(image).enhance(1.55)

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp:
        image.save(temp.name)
        temp_path = temp.name

    try:
        cmd = [exe, temp_path, "stdout"]
        if language:
            cmd.extend(["-l", language])
        cmd.extend(["--oem", "3", "--psm", "6", "--dpi", "300", "-c", "preserve_interword_spaces=1", "tsv"])

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=90,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
        )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or "Tesseract no pudo leer la imagen.")
        return result.stdout, image.size
    finally:
        try:
            os.unlink(temp_path)
        except OSError:
            pass


def run_tesseract_tsv_policia(image):
    """OCR especial para la planilla PNP.

    El formato de POLICÍA viene como captura de Excel con líneas de cuadrícula.
    En este tipo de imagen, el preprocesado fuerte + PSM 6 mezcla las celdas;
    PSM 3 sobre la captura casi original conserva mucho mejor cada fila.
    """
    exe = find_tesseract()
    if not exe:
        raise RuntimeError("Tesseract OCR no está instalado. Ejecuta instalar_ocr.bat.")

    languages = available_languages(exe)
    language = "spa" if "spa" in languages else ("eng" if "eng" in languages else None)

    # Mantener la cuadrícula y el texto tal como aparecen en Excel. Solo se amplían
    # capturas realmente pequeñas para que Tesseract no pierda caracteres.
    working = image.convert("RGB")
    if working.width < 900:
        ratio = 900 / working.width
        working = working.resize(
            (900, int(working.height * ratio)),
            Image.Resampling.LANCZOS,
        )

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp:
        working.save(temp.name)
        temp_path = temp.name

    try:
        cmd = [exe, temp_path, "stdout"]
        if language:
            cmd.extend(["-l", language])
        cmd.extend([
            "--oem", "3",
            "--psm", "3",
            "--dpi", "300",
            "-c", "preserve_interword_spaces=1",
            "tsv",
        ])
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=90,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
        )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or "Tesseract no pudo leer la imagen de POLICÍA.")
        return result.stdout, working.size
    finally:
        try:
            os.unlink(temp_path)
        except OSError:
            pass


def parse_words(tsv, preserve_accents=False):
    words = []
    reader = csv.DictReader(io.StringIO(tsv), delimiter="\t")
    for row in reader:
        text = clean_text(row.get("text"), preserve_accents=preserve_accents)
        try:
            confidence = float(row.get("conf", "-1"))
            x = int(row.get("left", "0"))
            y = int(row.get("top", "0"))
            w = int(row.get("width", "0"))
            h = int(row.get("height", "0"))
        except (ValueError, TypeError):
            continue
        if not text or confidence < 0:
            continue
        words.append({
            "text": text,
            "x": x,
            "right": x + w,
            "cx": x + w / 2,
            "cy": y + h / 2,
            "conf": confidence,
        })
    return words


def cluster_rows(words, image_height):
    tolerance = max(11, image_height * 0.012)
    rows = []
    for word in sorted(words, key=lambda item: (item["cy"], item["x"])):
        best = None
        best_distance = None
        for row in rows[-4:]:
            distance = abs(word["cy"] - row["center"])
            if distance <= tolerance and (best_distance is None or distance < best_distance):
                best = row
                best_distance = distance
        if best is None:
            rows.append({"center": word["cy"], "words": [word]})
        else:
            best["words"].append(word)
            best["center"] = statistics.mean(item["cy"] for item in best["words"])
    return sorted(rows, key=lambda row: row["center"])


def area_value(text):
    compact = re.sub(r"[^A-Z]", "", clean_text(text).upper())
    aliases = {
        "RAMPA": "RAMPA", "RANPA": "RAMPA", "RAMFA": "RAMPA", "RAMP": "RAMPA",
        "PAX": "PAX", "FAX": "PAX",
        "CARGA": "CARGA", "CARCA": "CARGA", "CARG": "CARGA",
        "OMA": "OMA",
    }
    return aliases.get(compact, "")


def identify_area_word(words, code_index):
    for index in range(code_index + 1, len(words)):
        value = area_value(words[index]["text"])
        if value:
            return index, value
    return None, ""


def extract_rows_talma(image_path, entry_start, second_start, observation_start):
    image = Image.open(image_path).convert("RGB")
    tsv, processed_size = run_tesseract_tsv(image)
    processed_width, processed_height = processed_size
    words = parse_words(tsv)
    grouped = cluster_rows(words, processed_height)

    if not (0.0 < entry_start < second_start < observation_start < 1.0):
        raise RuntimeError("Las líneas de columnas no están ordenadas correctamente.")

    output = []
    for group in grouped:
        row_words = sorted(group["words"], key=lambda item: item["x"])

        code_index = None
        code = ""
        for index, word in enumerate(row_words):
            digits = re.sub(r"\D", "", word["text"])
            if 7 <= len(digits) <= 9:
                position = word["cx"] / processed_width
                if 0.17 <= position <= 0.38:
                    code_index = index
                    code = digits
                    break
        if code_index is None:
            continue

        area_index, area = identify_area_word(row_words, code_index)
        if area_index is None:
            area_index = len(row_words)

        def joined(items):
            return clean_text(" ".join(item["text"] for item in items))

        code_x = row_words[code_index]["cx"] / processed_width
        before_code = row_words[:code_index]
        dates = [w for w in before_code if re.search(r"\d{1,2}/\d{1,2}/\d{4}", w["text"])]
        registro_words = []
        fecha_words = []
        sede_words = []
        if dates:
            first_date_index = before_code.index(dates[0])
            registro_words = before_code[:first_date_index + 1]
            remaining = before_code[first_date_index + 1:]
            if remaining and re.search(r"\d{1,2}/\d{1,2}/\d{4}", remaining[0]["text"]):
                fecha_words = [remaining[0]]
                sede_words = remaining[1:]
            else:
                fecha_words = dates[1:2]
                if len(dates) > 1:
                    second_idx = before_code.index(dates[1])
                    sede_words = before_code[second_idx + 1:]
                else:
                    sede_words = remaining
        else:
            sede_words = before_code

        name_words = row_words[code_index + 1:area_index]
        after_area = row_words[area_index + 1:] if area_index < len(row_words) else []

        entry_words = [
            w for w in after_area
            if entry_start <= w["cx"] / processed_width < second_start
        ]
        second_words = [
            w for w in after_area
            if second_start <= w["cx"] / processed_width < observation_start
        ]
        observation_words = [
            w for w in after_area
            if w["cx"] / processed_width >= observation_start
        ]

        row = {
            "registro": joined(registro_words),
            "fecha": joined(fecha_words),
            "sede": joined(sede_words),
            "codigo": code,
            "nombre": joined(name_words),
            "area": area,
            "entrada": joined(entry_words),
            "segundo": joined(second_words),
            "observacion": joined(observation_words),
        }

        warnings = []
        if not row["nombre"]:
            warnings.append("Nombre vacío")
        if not row["area"]:
            warnings.append("Área no reconocida")
        if not row["entrada"]:
            warnings.append("Entrada vacía: revisa la línea roja Entrada/Segundo")
        if not row["segundo"]:
            warnings.append("Segundo vacío: revisa la línea roja Segundo/Observación")
        if area_value(row["entrada"]):
            warnings.append("Entrada contiene un área")
        if area_value(row["segundo"]):
            warnings.append("Segundo contiene un área")
        row["_warnings"] = warnings
        output.append(row)

    if not output:
        raise RuntimeError("No se detectaron filas. Verifica que la captura incluya la columna Código y que el texto sea legible.")
    return output


HEADER_HINTS = re.compile(
    r"NOMBRE|APELLIDO|ENTRADA|PLATO|FONDO|SUGEREN|OBSERV|SEGUNDO|CODIGO|ÁREA|AREA|SEDE|REGISTRO|FECHA|^N[°ºoO]?$",
    re.IGNORECASE,
)


def normalize_policia_name(value):
    """Corrige errores OCR muy comunes en los grados PNP sin tocar apellidos/nombres."""
    text = clean_text_unicode(value)
    text = re.sub(r"(?i)^TNTE\s*[.\-]?\s*PNP\b", "TNTE PNP", text)
    text = re.sub(r"(?i)^ST1\s*[.\-]?\s*PNP\b", "ST1 PNP", text)
    text = re.sub(r"(?i)^ST3\s*[.\-]?\s*PNP\b", "ST3 PNP", text)
    # En esta planilla S2 suele ser leído como 2, 52, $2, s2 o S2.PNP.
    text = re.sub(r"(?i)^(?:S2|52|2)\s*[.\-]?\s*PNP\b", "S2 PNP", text)
    text = re.sub(r"(?i)\bPNP\b", "PNP", text)
    return clean_text_unicode(text)


def _match_case(original, replacement):
    if original.isupper():
        return replacement.upper()
    if original[:1].isupper():
        return replacement[:1].upper() + replacement[1:]
    return replacement


def normalize_policia_menu_text(value):
    """Conserva/restaura tildes frecuentes que OCR inglés suele perder o leer como 'e'."""
    text = clean_text_unicode(value)
    # Con Tesseract sin el paquete SPA, una ó pequeña suele reconocerse como 'e'.
    # Las sustituciones se limitan a vocabulario de comida para no alterar nombres.
    accent_fixes = {
        "limon": "limón", "limen": "limón",
        "lechon": "lechón", "lechen": "lechón",
        "salpicon": "salpicón", "salpicen": "salpicón",
        "pure": "puré",
        "atun": "atún", "jamon": "jamón", "camaron": "camarón",
        "chicharron": "chicharrón", "tallarin": "tallarín",
        "brocoli": "brócoli", "platano": "plátano", "maiz": "maíz",
        "aji": "ají", "higado": "hígado", "oregano": "orégano",
        "albondiga": "albóndiga", "albondigas": "albóndigas",
        "menesron": "menesrón", "menestron": "menestrón",
        "menu": "menú", "porcion": "porción", "guarnicion": "guarnición",
    }
    for wrong, right in accent_fixes.items():
        pattern = re.compile(rf"(?i)(?<![A-Za-zÁÉÍÓÚÜÑáéíóúüñ]){re.escape(wrong)}(?![A-Za-zÁÉÍÓÚÜÑáéíóúüñ])")
        text = pattern.sub(lambda m: _match_case(m.group(0), right), text)
    return clean_text_unicode(text)


def extract_rows_policia(image_path, entry_start, second_start, observation_start):
    """Extrae la planilla PNP: N° | Nombre | Entrada | Plato de fondo | Sugerencias.

    A diferencia de TALMA, la captura PNP incluye los números de fila de Excel a la
    izquierda. Se usa un OCR específico y se toma como N° el número dentro de la
    primera columna de la tabla, ignorando el número de fila de Excel.
    """
    image = Image.open(image_path).convert("RGB")
    tsv, processed_size = run_tesseract_tsv_policia(image)
    processed_width, processed_height = processed_size
    words = parse_words(tsv, preserve_accents=True)
    grouped = cluster_rows(words, processed_height)

    if not (0.0 < entry_start < second_start < observation_start < 1.0):
        raise RuntimeError("Las líneas de columnas no están ordenadas correctamente.")

    output = []
    for group in grouped:
        row_words = sorted(group["words"], key=lambda item: item["x"])
        if not row_words:
            continue

        def joined(items):
            return clean_text_unicode(" ".join(item["text"] for item in items))

        # Todo lo anterior a ENTRADA contiene: [número de fila Excel] [N°] [nombre].
        left_words = [w for w in row_words if w["cx"] / processed_width < entry_start]

        # El N° real está dentro de la tabla. El número de fila de Excel queda mucho
        # más pegado al borde izquierdo y se excluye con este margen. Elegimos el
        # primer número corto de la zona N°; así tampoco confundimos "S2" con N°.
        number_candidates = []
        number_min_x = max(0.03, entry_start * 0.09)
        number_max_x = min(entry_start * 0.30, entry_start - 0.02)
        for index, word in enumerate(left_words):
            digits = re.sub(r"\D", "", word["text"])
            position = word["cx"] / processed_width
            if digits and len(digits) <= 3 and number_min_x <= position <= number_max_x:
                number_candidates.append((index, word, digits))

        if not number_candidates:
            continue

        number_index, _, numero = number_candidates[0]
        name_items = left_words[number_index + 1:]
        nombre = normalize_policia_name(joined(name_items))

        # Una fila PNP válida siempre debe tener un nombre razonable. Esto descarta
        # encabezados, fecha, letras de columnas de Excel y filas vacías.
        if not nombre or len(re.sub(r"[^A-Za-zÁÉÍÓÚÜÑáéíóúüñ]", "", nombre)) < 4:
            continue

        entry_words = [
            w for w in row_words
            if entry_start <= w["cx"] / processed_width < second_start
        ]
        second_words = [
            w for w in row_words
            if second_start <= w["cx"] / processed_width < observation_start
        ]
        observation_words = [
            w for w in row_words
            if w["cx"] / processed_width >= observation_start
        ]

        row = {
            "numero": numero,
            "nombre": nombre,
            "entrada": normalize_policia_menu_text(joined(entry_words)),
            "segundo": normalize_policia_menu_text(joined(second_words)),
            "observacion": normalize_policia_menu_text(joined(observation_words)),
        }

        warnings = []
        if not row["entrada"]:
            warnings.append("Entrada vacía: revisa la línea de Entrada")
        if not row["segundo"]:
            warnings.append("Plato de fondo vacío: revisa la línea de Plato de fondo")
        row["_warnings"] = warnings
        output.append(row)

    if not output:
        raise RuntimeError(
            "No se detectaron filas de POLICÍA. Verifica que la captura tenga el formato "
            "N°, Nombre y apellido, Entrada, Plato de fondo y Sugerencias, y ajusta las barras."
        )
    return output


def extract_rows(sistema, image_path, entry_start, second_start, observation_start):
    if sistema == "policia":
        return extract_rows_policia(image_path, entry_start, second_start, observation_start)
    return extract_rows_talma(image_path, entry_start, second_start, observation_start)


def save_rows(sistema, rows):
    path = cfg(sistema)["data_file"]
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def load_rows(sistema):
    path = cfg(sistema)["data_file"]
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []


@app.errorhandler(Exception)
def error_handler(error):
    detail = traceback.format_exc()
    try:
        return render_template("error.html", message=str(error), detail=detail), 500
    except Exception:
        # Último recurso para que un fallo de plantilla nunca termine como 502 opaco.
        return Response(
            "<!doctype html><html lang='es'><meta charset='utf-8'><title>Error</title>"
            "<body style='font-family:Arial;padding:24px'><h1>Error del panel</h1><p>"
            + str(escape(str(error)))
            + "</p><pre style='white-space:pre-wrap'>"
            + str(escape(detail))
            + "</pre><p><a href='/admin/dashboard'>Volver al panel</a></p></body></html>",
            status=500,
            mimetype="text/html",
        )


@app.get("/admin/scanner-style.css")
def scanner_style():
    return Response(SCANNER_CSS, mimetype="text/css")


@app.get("/admin/scanners")
def index():
    return render_template("index.html")


@app.get("/admin/<any(talma,policia):sistema>/")
def system_home(sistema):
    c = cfg(sistema)
    return render_template(
        "subir.html",
        sistema=sistema,
        titulo=c["titulo"],
        subtitulo=c["subtitulo"],
        ayuda=c["ayuda"],
    )


@app.post("/admin/<any(talma,policia):sistema>/subir")
def upload_image(sistema):
    cfg(sistema)  # valida
    uploaded = request.files.get("imagen")
    if not uploaded or not uploaded.filename:
        raise RuntimeError("Selecciona una imagen.")
    suffix = Path(uploaded.filename).suffix.lower()
    if suffix not in (".png", ".jpg", ".jpeg", ".webp"):
        raise RuntimeError("La imagen debe ser PNG, JPG, JPEG o WEBP.")
    filename = f"{uuid.uuid4().hex}{suffix}"
    image_path = UPLOADS / filename
    uploaded.save(image_path)
    return redirect(url_for("calibrate", sistema=sistema, filename=filename))


@app.get("/admin/<any(talma,policia):sistema>/calibrar/<filename>")
def calibrate(sistema, filename):
    c = cfg(sistema)
    image_path = UPLOADS / Path(filename).name
    if not image_path.exists():
        raise RuntimeError("La imagen ya no existe. Vuelve a cargarla.")
    return render_template(
        "calibrar.html",
        sistema=sistema,
        titulo=c["titulo"],
        filename=image_path.name,
        entry_start=c["default_entry"],
        second_start=c["default_second"],
        observation_start=c["default_observation"],
        label_entrada=c["label_entrada"],
        label_segundo=c["label_segundo"],
        label_observacion=c["label_observacion"],
    )


@app.get("/admin/scanner-imagen/<filename>")
def uploaded_image(filename):
    image_path = UPLOADS / Path(filename).name
    if not image_path.exists():
        raise RuntimeError("Imagen no encontrada.")
    return send_file(image_path)


@app.post("/admin/<any(talma,policia):sistema>/procesar")
def process_image(sistema):
    c = cfg(sistema)
    filename = Path(request.form.get("filename", "")).name
    image_path = UPLOADS / filename
    if not filename or not image_path.exists():
        raise RuntimeError("La imagen no existe. Vuelve a cargarla.")
    try:
        entry_start = float(request.form.get("entry_start", c["default_entry"]))
        second_start = float(request.form.get("second_start", c["default_second"]))
        observation_start = float(request.form.get("observation_start", c["default_observation"]))
    except ValueError:
        raise RuntimeError("Los límites de columnas no son válidos.")

    rows = extract_rows(sistema, image_path, entry_start, second_start, observation_start)
    save_rows(sistema, rows)

    if sistema == "policia":
        return render_template("revisar_policia.html", rows=rows, filename=filename)
    return render_template(
        "revisar_talma.html",
        rows=rows,
        areas=c["areas"],
        filename=filename,
    )


@app.post("/admin/<any(talma,policia):sistema>/guardar")
def save_changes(sistema):
    c = cfg(sistema)
    received = json.loads(request.form.get("rows_json", "[]"))
    cleaned = []
    for row in received:
        cleaner = clean_text_unicode if sistema == "policia" else clean_text
        item = {column: cleaner(row.get(column, "")) for column in c["cols"]}
        if sistema == "talma":
            item["codigo"] = re.sub(r"\D", "", item.get("codigo", ""))
            item["area"] = area_value(item.get("area", ""))
        if sistema == "policia":
            item["numero"] = re.sub(r"\D", "", item.get("numero", ""))
            for field in ("entrada", "segundo", "observacion"):
                item[field] = normalize_policia_menu_text(item.get(field, ""))
        item["_warnings"] = []
        cleaned.append(item)
    save_rows(sistema, cleaned)
    return jsonify({"ok": True, "count": len(cleaned)})


def draw_centered_fit(pdf, text, center_x, y, max_width, font="Helvetica-Bold", max_size=16, min_size=9, preserve_accents=False):
    safe_text = pdf_text(text, preserve_accents=preserve_accents)
    size = max_size
    while size > min_size and pdf.stringWidth(safe_text, font, size) > max_width:
        size -= 1
    if pdf.stringWidth(safe_text, font, size) > max_width:
        original = safe_text
        while len(safe_text) > 4 and pdf.stringWidth(safe_text + "...", font, size) > max_width:
            safe_text = safe_text[:-1].rstrip()
        if safe_text != original:
            safe_text += "..."
    pdf.setFont(font, size)
    pdf.drawCentredString(center_x, y, safe_text)


def _wrap_lines_for_size(pdf, text, max_width, font, size, preserve_accents=False):
    words = pdf_text(text, preserve_accents=preserve_accents).split()
    if not words:
        return []
    lines = []
    current = ""
    for word in words:
        proposed = f"{current} {word}".strip()
        if pdf.stringWidth(proposed, font, size) <= max_width:
            current = proposed
            continue
        if current:
            lines.append(current)
            current = ""
        piece = ""
        for char in word:
            candidate = piece + char
            if piece and pdf.stringWidth(candidate, font, size) > max_width:
                lines.append(piece)
                piece = char
            else:
                piece = candidate
        current = piece
    if current:
        lines.append(current)
    return lines


def draw_text_in_box(
    pdf,
    text,
    x,
    top_y,
    max_width,
    max_height,
    font="Helvetica-Bold",
    max_size=18,
    min_size=10,
    max_lines=3,
    preserve_accents=False,
):
    safe_text = pdf_text(text, preserve_accents=preserve_accents)
    if not safe_text:
        return
    chosen_size = min_size
    chosen_lines = [safe_text]
    chosen_leading = min_size * 1.08
    for size in range(max_size, min_size - 1, -1):
        lines = _wrap_lines_for_size(pdf, safe_text, max_width, font, size, preserve_accents=preserve_accents)
        leading = size * 1.08
        required_height = size + max(0, len(lines) - 1) * leading
        if len(lines) <= max_lines and required_height <= max_height:
            chosen_size = size
            chosen_lines = lines
            chosen_leading = leading
            break
    if len(chosen_lines) > max_lines:
        chosen_lines = chosen_lines[:max_lines]
        last = chosen_lines[-1]
        suffix = "..."
        while last and pdf.stringWidth(last + suffix, font, chosen_size) > max_width:
            last = last[:-1].rstrip()
        chosen_lines[-1] = (last + suffix) if last else suffix
    pdf.setFont(font, chosen_size)
    baseline = top_y - chosen_size
    for line in chosen_lines:
        pdf.drawString(x, baseline, line)
        baseline -= chosen_leading


def count_order_values(rows, field, preserve_accents=False):
    counts = Counter()
    labels = {}
    for row in rows:
        value = clean_text(row.get(field, ""), preserve_accents=preserve_accents)
        if not value:
            continue
        key = value.casefold()
        counts[key] += 1
        labels.setdefault(key, value)
    return sorted(
        ((labels[key], amount) for key, amount in counts.items()),
        key=lambda item: (-item[1], item[0].casefold()),
    )


def fit_single_line(pdf, text, max_width, font="Helvetica", size=10, preserve_accents=False):
    value = pdf_text(text, preserve_accents=preserve_accents)
    if pdf.stringWidth(value, font, size) <= max_width:
        return value
    suffix = "..."
    while value and pdf.stringWidth(value + suffix, font, size) > max_width:
        value = value[:-1].rstrip()
    return value + suffix if value else suffix


def draw_pdf_summary_talma(pdf, rows, page_width, page_height, areas):
    left = 16 * mm
    right = page_width - 16 * mm
    top = page_height - 16 * mm
    generated_at = datetime.now()

    area_counts = Counter()
    for row in rows:
        area = clean_text(row.get("area", "")).upper() or "SIN ÁREA"
        area_counts[area] += 1

    ordered_areas = []
    for area in areas:
        if area_counts.get(area):
            ordered_areas.append((area, area_counts.pop(area)))
    ordered_areas.extend(sorted(area_counts.items(), key=lambda item: (-item[1], item[0])))

    pdf.showPage()
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(left, top, "RESUMEN DE PEDIDOS — TALMA")
    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(right, top + 1, generated_at.strftime("%d/%m/%Y %H:%M"))
    pdf.setLineWidth(0.8)
    pdf.line(left, top - 4 * mm, right, top - 4 * mm)

    y = top - 14 * mm
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(left, y, f"TOTAL DE PEDIDOS: {len(rows)}")
    y -= 13 * mm

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(left, y, "PEDIDOS POR ÁREA / CATEGORÍA")
    pdf.setLineWidth(0.4)
    pdf.line(left, y - 2 * mm, right, y - 2 * mm)
    y -= 9 * mm

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left + 2 * mm, y, "ÁREA / CATEGORÍA")
    pdf.drawRightString(right - 2 * mm, y, "CANTIDAD")
    y -= 7 * mm

    if not ordered_areas:
        pdf.setFont("Helvetica-Oblique", 10)
        pdf.drawString(left + 2 * mm, y, "SIN DATOS")
        return

    for area, amount in ordered_areas:
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(left + 2 * mm, y, pdf_text(area))
        pdf.drawRightString(right - 2 * mm, y, str(amount))
        pdf.setLineWidth(0.25)
        pdf.line(left + 2 * mm, y - 2 * mm, right - 2 * mm, y - 2 * mm)
        y -= 8 * mm


def draw_pdf_summary_policia(pdf, rows, page_width, page_height):
    left = 16 * mm
    right = page_width - 16 * mm
    top = page_height - 16 * mm
    generated_at = datetime.now()

    pdf.showPage()
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(left, top, "RESUMEN DE PEDIDOS — POLICÍA")
    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(right, top + 1, generated_at.strftime("%d/%m/%Y %H:%M"))
    pdf.setLineWidth(0.8)
    pdf.line(left, top - 4 * mm, right, top - 4 * mm)

    y = top - 14 * mm
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(left, y, f"TOTAL DE PEDIDOS: {len(rows)}")
    y -= 14 * mm

    for field, title in (
        ("entrada", "ENTRADAS MÁS PEDIDAS"),
        ("segundo", "PLATOS DE FONDO MÁS PEDIDOS"),
        ("observacion", "SUGERENCIAS"),
    ):
        values = count_order_values(rows, field, preserve_accents=True)
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawString(left, y, title)
        pdf.setLineWidth(0.4)
        pdf.line(left, y - 2 * mm, right, y - 2 * mm)
        y -= 8 * mm
        if not values:
            pdf.setFont("Helvetica-Oblique", 10)
            pdf.drawString(left + 2 * mm, y, "SIN DATOS")
            y -= 10 * mm
            continue
        for label, amount in values[:12]:
            if y < 25 * mm:
                pdf.showPage()
                y = top
            pdf.setFont("Helvetica", 11)
            pdf.drawString(left + 2 * mm, y, fit_single_line(pdf, label, right - left - 30 * mm, size=11, preserve_accents=True))
            pdf.drawRightString(right - 2 * mm, y, str(amount))
            y -= 6.5 * mm
        y -= 6 * mm


@app.get("/admin/<any(talma,policia):sistema>/pdf")
def create_pdf(sistema):
    c = cfg(sistema)
    rows = load_rows(sistema)
    if not rows:
        raise RuntimeError("No hay pedidos guardados para generar el PDF.")
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    output = OUTPUTS / f"cuponera_{c['pdf_prefix']}_{datetime.now():%Y%m%d_%H%M%S}.pdf"

    pdf = canvas.Canvas(str(output), pagesize=A4)
    page_width, page_height = A4
    preserve_accents = sistema == "policia"
    margin = 8 * mm
    coupon_width = (page_width - 2 * margin) / 2
    coupon_height = (page_height - 2 * margin) / 4

    for index, row in enumerate(rows):
        position = index % 8
        if index and position == 0:
            pdf.showPage()
        column = position % 2
        line = position // 2
        x = margin + column * coupon_width
        top = page_height - margin - line * coupon_height
        pad = 5 * mm

        pdf.rect(x, top - coupon_height, coupon_width, coupon_height)
        text_x = x + pad
        text_y = top - 7 * mm

        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(text_x, text_y, f"TICKET {index + 1:03d}")

        if sistema == "talma":
            header = pdf_text(row.get("area", "") or "SIN ÁREA").upper()
            fecha = pdf_text(row.get("fecha", ""))
            header_size = 12
            name_max_size = 12
            name_min_size = 8
            content_top = top - 19 * mm
            first_box_height = 16 * mm
            second_box_top = content_top - 19 * mm
            second_box_height = 16 * mm
            menu_max_size = 35
            menu_min_size = 17
            observation_label_y = content_top - 39 * mm
        else:
            header = "POLICÍA"
            fecha = ""
            header_size = 12
            name_max_size = 13
            name_min_size = 8
            # POLICÍA: empresa en esquina; nombre/área más discretos y platos mucho más grandes.
            content_top = top - 19 * mm
            first_box_height = 17 * mm
            second_box_top = content_top - 20 * mm
            second_box_height = 17 * mm
            menu_max_size = 39
            menu_min_size = 18
            observation_label_y = content_top - 42 * mm
            num = clean_text(row.get("numero", ""))
            if num:
                pdf.setFont("Helvetica", 8)
                pdf.drawString(text_x + 22 * mm, text_y, f"N° {num}")

        # Identificador de empresa: pequeño, visible y siempre en la esquina superior derecha.
        company_label = "TALMA" if sistema == "talma" else "POLICÍA"
        label_w = 16 * mm
        label_h = 4.5 * mm
        label_x = x + coupon_width - pad - label_w
        label_y = top - label_h - 2.5 * mm
        pdf.setLineWidth(0.5)
        pdf.rect(label_x, label_y, label_w, label_h)
        pdf.setFont("Helvetica-Bold", 7)
        pdf.drawCentredString(label_x + label_w / 2, label_y + 1.35 * mm, company_label)

        pdf.setFont("Helvetica-Bold", header_size)
        pdf.drawCentredString(x + coupon_width / 2, text_y, header)
        if fecha:
            pdf.setFont("Helvetica", 7)
            pdf.drawRightString(x + coupon_width - pad, text_y, fecha)
        text_y -= 7 * mm

        name = pdf_text(row.get("nombre", ""), preserve_accents=preserve_accents).upper()
        draw_centered_fit(
            pdf,
            name,
            x + coupon_width / 2,
            text_y,
            coupon_width - 2 * pad,
            font="Helvetica-Bold",
            max_size=name_max_size,
            min_size=name_min_size,
            preserve_accents=preserve_accents,
        )

        ticket_bottom = top - coupon_height + 4 * mm
        inner_width = coupon_width - 2 * pad

        draw_text_in_box(
            pdf,
            row.get("entrada", ""),
            text_x,
            content_top,
            inner_width,
            first_box_height,
            font="Helvetica-Bold",
            max_size=menu_max_size,
            min_size=menu_min_size,
            max_lines=2,
            preserve_accents=preserve_accents,
        )

        draw_text_in_box(
            pdf,
            row.get("segundo", ""),
            text_x,
            second_box_top,
            inner_width,
            second_box_height,
            font="Helvetica-Bold",
            max_size=menu_max_size,
            min_size=menu_min_size,
            max_lines=2,
            preserve_accents=preserve_accents,
        )

        obs_label = "OBSERVACIÓN" if sistema == "talma" else "SUGERENCIAS"
        if clean_text(row.get("observacion", "")):
            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(text_x, observation_label_y, obs_label)
            observation_top = observation_label_y - 2 * mm
            observation_height = max(7 * mm, observation_top - ticket_bottom)
            draw_text_in_box(
                pdf,
                row.get("observacion", ""),
                text_x,
                observation_top,
                inner_width,
                observation_height,
                font="Helvetica-Bold",
                max_size=10,
                min_size=7,
                max_lines=3,
                preserve_accents=preserve_accents,
            )

    if sistema == "talma":
        draw_pdf_summary_talma(pdf, rows, page_width, page_height, c["areas"])
    else:
        draw_pdf_summary_policia(pdf, rows, page_width, page_height)

    pdf.save()
    return send_file(output, as_attachment=True)


# ---- Excel helpers (compartidos, parametrizados por sistema) ----

def normalized_excel_header(values):
    return [clean_text(str(v or "")).casefold() for v in values]


def is_orders_sheet(sheet, headers):
    expected = [clean_text(str(h or "")).casefold() for h in headers]
    first = normalized_excel_header([cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))])
    return first[: len(expected)] == expected


def parse_order_date(value):
    text = clean_text(str(value or ""))
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def order_date_label(rows, date_field="fecha"):
    dates = []
    for row in rows:
        parsed = parse_order_date(row.get(date_field, ""))
        if parsed:
            dates.append(parsed)
    if not dates:
        return datetime.now().strftime("%d/%m/%Y")
    lo, hi = min(dates), max(dates)
    if lo.date() == hi.date():
        return lo.strftime("%d/%m/%Y")
    return f"{lo.strftime('%d/%m/%Y')} - {hi.strftime('%d/%m/%Y')}"


def sheet_date_range(sheet, date_col_index=2):
    dates = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or date_col_index >= len(row):
            continue
        parsed = parse_order_date(row[date_col_index])
        if parsed:
            dates.append(parsed)
    if not dates:
        return ""
    lo, hi = min(dates), max(dates)
    if lo.date() == hi.date():
        return lo.strftime("%d/%m/%Y")
    return f"{lo.strftime('%d/%m/%Y')} - {hi.strftime('%d/%m/%Y')}"


def style_orders_sheet(sheet, widths):
    header_fill = PatternFill("solid", fgColor="176B43")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(widths)):
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def remove_previous_summary(sheet):
    # Elimina resúmenes anteriores para que la actualización mensual no duplique totales.
    start = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        text = " ".join(clean_text(str(v or "")) for v in row).upper()
        if "RESUMEN DE ALMUERZOS POR PERSONA" in text:
            start = idx
            break
    if start is not None:
        sheet.delete_rows(start, sheet.max_row - start + 1)

    rows_to_delete = []
    for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        text = " ".join(clean_text(str(v or "")) for v in row).upper()
        if "TOTAL DE PEDIDOS" in text or "PERIODO" in text or "PERÍODO" in text:
            rows_to_delete.append(idx)
    for idx in reversed(rows_to_delete):
        sheet.delete_rows(idx, 1)


def append_summary(sheet, added_count, accumulated_total, cols_count, period_label=""):
    sheet.append([])
    sheet.append(["PERIODO", period_label or datetime.now().strftime("%d/%m/%Y")])
    sheet.append(["PEDIDOS AGREGADOS", added_count])
    sheet.append(["TOTAL DE PEDIDOS ACUMULADO", accumulated_total])
    for row_idx in range(sheet.max_row - 2, sheet.max_row + 1):
        for col in range(1, min(3, cols_count + 1)):
            cell = sheet.cell(row=row_idx, column=col)
            cell.font = Font(bold=True)


def person_key(value):
    return clean_text(str(value or ""), preserve_accents=True).casefold()


def append_orders(sheet, rows, cols, preserve_accents=False):
    """Agrega pedidos y deja un número de almuerzo acumulativo por persona."""
    # Buscar el nombre dentro de las columnas del sistema.
    name_index = cols.index("nombre") if "nombre" in cols else None
    existing_counts = Counter()
    if name_index is not None:
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not values or name_index >= len(values):
                continue
            key = person_key(values[name_index])
            if key:
                existing_counts[key] += 1

    for row in rows:
        values = [clean_text(row.get(column, ""), preserve_accents=preserve_accents) for column in cols]
        if name_index is not None:
            key = person_key(row.get("nombre", ""))
            existing_counts[key] += 1
            values.insert(name_index + 1, existing_counts[key])
        sheet.append(values)


def recompute_lunch_numbers(sheet, headers):
    """Recalcula N° de almuerzo por persona en el orden en que aparecen los pedidos."""
    if "N° Almuerzo" not in headers:
        return
    name_col = headers.index("Nombre") + 1 if "Nombre" in headers else (
        headers.index("Nombre y apellido") + 1 if "Nombre y apellido" in headers else None
    )
    num_col = headers.index("N° Almuerzo") + 1
    if not name_col:
        return
    counts = Counter()
    for row_idx in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row_idx, column=name_col).value
        key = person_key(name)
        if not key:
            continue
        counts[key] += 1
        sheet.cell(row=row_idx, column=num_col).value = counts[key]


def append_person_summary(sheet, headers):
    """Añade al final de la hoja el total acumulado de almuerzos por persona."""
    name_col = headers.index("Nombre") + 1 if "Nombre" in headers else (
        headers.index("Nombre y apellido") + 1 if "Nombre y apellido" in headers else None
    )
    if not name_col:
        return
    counts = Counter()
    labels = {}
    for row_idx in range(2, sheet.max_row + 1):
        value = clean_text_unicode(str(sheet.cell(row=row_idx, column=name_col).value or "")).strip()
        if not value:
            continue
        key = value.casefold()
        counts[key] += 1
        labels.setdefault(key, value)

    sheet.append([])
    sheet.append(["RESUMEN DE ALMUERZOS POR PERSONA"])
    sheet.append(["PERSONA", "TOTAL ALMUERZOS"])
    for key, total in sorted(counts.items(), key=lambda item: (-item[1], labels[item[0]].casefold())):
        sheet.append([labels[key], total])

    title_row = sheet.max_row - len(counts) - 1
    for cell in sheet[title_row]:
        cell.font = Font(bold=True, size=12)
    for cell in sheet[title_row + 1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="176B43")
        cell.alignment = Alignment(horizontal="center")


def count_order_rows(sheet, headers):
    count = 0
    expected_len = len(headers)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        values = [clean_text(str(v or "")) for v in row[:expected_len]]
        text = " ".join(values).upper()
        if "TOTAL DE PEDIDOS" in text or "PERIODO" in text or "PERÍODO" in text:
            continue
        if any(values):
            count += 1
    return count


def prepare_orders_sheet(workbook, headers, preserve_accents=False):
    # Preferir hoja llamada Pedidos; consolidar si hay varias.
    candidates = [s for s in workbook.worksheets if s.title.lower().startswith("pedido")]
    if not candidates:
        sheet = workbook.worksheets[0] if workbook.worksheets else workbook.create_sheet("Pedidos")
        sheet.title = "Pedidos"
        if sheet.max_row < 1 or not is_orders_sheet(sheet, headers):
            for _ in range(sheet.max_row):
                sheet.delete_rows(1)
            sheet.append(headers)
        return sheet

    main = candidates[0]
    main.title = "Pedidos"
    current = normalized_excel_header([cell.value for cell in next(main.iter_rows(min_row=1, max_row=1))]) if main.max_row else []
    expected = normalized_excel_header(headers)

    if current != expected:
        # Migración compatible de Excel mensual anterior: conserva pedidos y agrega N° Almuerzo.
        legacy = headers.copy()
        if "N° Almuerzo" in legacy:
            legacy.remove("N° Almuerzo")
        legacy_expected = normalized_excel_header(legacy)
        if current[:len(legacy_expected)] == legacy_expected:
            insert_at = legacy.index("Nombre") + 2 if "Nombre" in legacy else legacy.index("Nombre y apellido") + 2
            main.insert_cols(insert_at)
            main.cell(row=1, column=insert_at).value = "N° Almuerzo"
        else:
            for _ in range(main.max_row):
                main.delete_rows(1)
            main.append(headers)

    for extra in candidates[1:]:
        for row in extra.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            values = [clean_text(str(v or ""), preserve_accents=preserve_accents) for v in row[: len(headers)]]
            text = " ".join(values).upper()
            if "TOTAL DE PEDIDOS" in text or "PERIODO" in text or "RESUMEN DE ALMUERZOS" in text:
                continue
            if any(values):
                main.append(values)
        del workbook[extra.title]
    return main


@app.get("/admin/<any(talma,policia):sistema>/excel")
def create_excel(sistema):
    c = cfg(sistema)
    rows = load_rows(sistema)
    if not rows:
        raise RuntimeError("No hay pedidos guardados para exportar.")
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    output = OUTPUTS / f"pedidos_{c['pdf_prefix']}_{datetime.now():%Y_%m}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pedidos"
    sheet.append(c["excel_headers"])
    append_orders(sheet, rows, c["cols"], preserve_accents=(sistema == "policia"))
    recompute_lunch_numbers(sheet, c["excel_headers"])
    total = count_order_rows(sheet, c["excel_headers"])
    period = order_date_label(rows, "fecha") if sistema == "talma" else datetime.now().strftime("%d/%m/%Y")
    append_summary(sheet, added_count=len(rows), accumulated_total=total, cols_count=len(c["excel_headers"]), period_label=period)
    append_person_summary(sheet, c["excel_headers"])
    style_orders_sheet(sheet, c["excel_widths"])
    workbook.save(output)
    return send_file(output, as_attachment=True, download_name=output.name)


@app.post("/admin/<any(talma,policia):sistema>/actualizar-excel")
def update_excel(sistema):
    c = cfg(sistema)
    rows = load_rows(sistema)
    if not rows:
        raise RuntimeError("No hay pedidos guardados para agregar al Excel.")

    uploaded = request.files.get("excel_anterior")
    if not uploaded or not uploaded.filename:
        raise RuntimeError("Selecciona el archivo Excel mensual que deseas actualizar.")
    suffix = Path(uploaded.filename).suffix.lower()
    if suffix != ".xlsx":
        raise RuntimeError("El archivo anterior debe estar en formato .xlsx.")

    OUTPUTS.mkdir(parents=True, exist_ok=True)
    temp_input = UPLOADS / f"{uuid.uuid4().hex}.xlsx"
    uploaded.save(temp_input)
    try:
        workbook = load_workbook(temp_input)
        sheet = prepare_orders_sheet(workbook, c["excel_headers"], preserve_accents=(sistema == "policia"))
        remove_previous_summary(sheet)
        append_orders(sheet, rows, c["cols"], preserve_accents=(sistema == "policia"))
        recompute_lunch_numbers(sheet, c["excel_headers"])
        accumulated_total = count_order_rows(sheet, c["excel_headers"])
        period = order_date_label(rows, "fecha") if sistema == "talma" else datetime.now().strftime("%d/%m/%Y")
        append_summary(
            sheet,
            added_count=len(rows),
            accumulated_total=accumulated_total,
            cols_count=len(c["excel_headers"]),
            period_label=period,
        )
        append_person_summary(sheet, c["excel_headers"])
        style_orders_sheet(sheet, c["excel_widths"])

        original_name = clean_text(Path(uploaded.filename).stem) or f"pedidos_{c['pdf_prefix']}"
        output = OUTPUTS / f"{original_name}_actualizado.xlsx"
        workbook.save(output)
    finally:
        try:
            temp_input.unlink()
        except OSError:
            pass

    return send_file(output, as_attachment=True, download_name=output.name)


@app.get("/admin/scanner-diagnostico")
def diagnostic():
    exe = find_tesseract()
    return jsonify({
        "version": "20.3-talma-dni",
        "sistemas": list(SISTEMAS.keys()),
        "tesseract_encontrado": bool(exe),
        "ruta": exe,
        "idiomas": available_languages(exe) if exe else [],
        "python": os.sys.version,
        "uploads": str(UPLOADS),
        "outputs": str(OUTPUTS),
        "templates_dir": str(TEMPLATES_DIR),
        "templates_dir_existe": TEMPLATES_DIR.exists(),
        "templates_en_disco": sorted(p.name for p in TEMPLATES_DIR.glob("*.html")) if TEMPLATES_DIR.exists() else [],
        "templates_embebidas": sorted(SCANNER_TEMPLATES.keys()),
    })



# ---------------------------------------------------------------------------
# Portal privado TALMA
# ---------------------------------------------------------------------------
def _talma_password() -> str:
    return os.environ.get("TALMA_PASSWORD", "Talma2026")


def _talma_session_value() -> str:
    expires = str(int(time.time()) + 12 * 3600)
    payload = f"talma:{expires}"
    sig = hmac.new(CONFIG["secret_key"].encode(), payload.encode(), hashlib.sha256).hexdigest()
    return f"{payload}:{sig}"


def _valid_talma_session(value: str | None) -> bool:
    if not value:
        return False
    try:
        user, expires, sig = value.split(":", 2)
        if user != "talma" or int(expires) < time.time():
            return False
        payload = f"{user}:{expires}"
        expected = hmac.new(CONFIG["secret_key"].encode(), payload.encode(), hashlib.sha256).hexdigest()
        return hmac.compare_digest(sig, expected)
    except (ValueError, TypeError):
        return False


def _talma_page(title: str, body: str, extra: str = "") -> str:
    return page(title, f"""
<main class="wrap">
<div class="actions no-print" style="justify-content:space-between;margin-bottom:16px">
  <h1 style="margin:0">Portal TALMA</h1>
  <span class="actions"><a class="btn secondary" href="/talma/excel">Descargar Excel</a>
  <a class="btn secondary" href="/talma/logout">Cerrar sesión</a></span>
</div>
{body}
</main>
{extra}
""")


def _talma_login(error: bool = False):
    notice = '<div class="notice error">Contraseña incorrecta.</div>' if error else ""
    body = f"""
<main class="wrap narrow">
<div class="card">
<h1>Acceso TALMA</h1>
<p class="muted">Este apartado es exclusivo para TALMA.</p>
{notice}
<form method="post" action="/talma/login">
<label>Contraseña TALMA</label>
<input type="password" name="password" required autofocus autocomplete="current-password">
<button type="submit">Ingresar</button>
</form>
</div>
</main>"""
    return page("Acceso TALMA", body)


@app.get("/talma")
@app.get("/talma/")
def talma_portal():
    if not _valid_talma_session(request.cookies.get("talma_session")):
        return _talma_login()

    orders_app.init_db()
    dni_filter = re.sub(r"\D", "", request.args.get("dni", ""))[:8]
    fecha_desde = request.args.get("desde", "").strip()
    fecha_hasta = request.args.get("hasta", "").strip()

    # Si solo se indica una fecha, se usa como fecha única.
    if fecha_desde and not fecha_hasta:
        fecha_hasta = fecha_desde
    if fecha_hasta and not fecha_desde:
        fecha_desde = fecha_hasta

    def valid_iso(value):
        try:
            datetime.strptime(value, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    if fecha_desde and not valid_iso(fecha_desde):
        fecha_desde = ""
    if fecha_hasta and not valid_iso(fecha_hasta):
        fecha_hasta = ""
    if fecha_desde and fecha_hasta and fecha_desde > fecha_hasta:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    with orders_app.db() as conn:
        sql = """SELECT o.id, o.order_date, o.employee_name, o.dni, o.area, o.entry_item,
                        o.main_item, o.notes, o.created_at
                 FROM orders o JOIN companies c ON c.id=o.company_id
                 WHERE c.slug='talma'"""
        args = []
        if dni_filter:
            sql += " AND o.dni=?"
            args.append(dni_filter)
        if fecha_desde:
            sql += " AND o.order_date>=?"
            args.append(fecha_desde)
        if fecha_hasta:
            sql += " AND o.order_date<=?"
            args.append(fecha_hasta)
        sql += """ ORDER BY o.order_date DESC,
                   CASE WHEN o.dni='' THEN 1 ELSE 0 END,
                   o.dni, o.employee_name COLLATE NOCASE, o.id"""
        rows = conn.execute(sql, args).fetchall()

    counts = {}
    for r in rows:
        key = r["dni"] or f"legacy:{orders_app.normalize_key(r['employee_name'])}"
        counts[key] = counts.get(key, 0) + 1

    running = {}
    table_rows = []
    for r in rows:
        key = r["dni"] or f"legacy:{orders_app.normalize_key(r['employee_name'])}"
        running[key] = running.get(key, 0) + 1
        table_rows.append(
            f"<tr><td>{esc(r['order_date'])}</td><td><b>{esc(r['dni'] or '—')}</b></td><td>{esc(r['employee_name'])}</td>"
            f"<td>{esc(r['area'])}</td><td>{esc(r['entry_item'])}</td><td>{esc(r['main_item'])}</td>"
            f"<td>{running[key]}</td><td>{esc((r['created_at'] or '')[11:16])}</td></tr>"
        )

    summary = []
    for dni_key, total in sorted(counts.items(), key=lambda x: x[0]):
        rr = next((r for r in rows if (r["dni"] or f"legacy:{orders_app.normalize_key(r['employee_name'])}") == dni_key), None)
        label = rr["employee_name"] if rr else dni_key
        dni_label = rr["dni"] if rr and rr["dni"] else "Sin DNI (registro antiguo)"
        summary.append(
            f"<tr><td><b>{esc(dni_label)}</b></td><td>{esc(label)}</td><td>{total}</td></tr>"
        )

    if fecha_desde and fecha_hasta:
        periodo_label = (
            f"{esc(orders_app.fecha_con_dia(fecha_desde))}"
            if fecha_desde == fecha_hasta
            else f"del <b>{esc(orders_app.fecha_con_dia(fecha_desde))}</b> "
                 f"al <b>{esc(orders_app.fecha_con_dia(fecha_hasta))}</b>"
        )
    else:
        periodo_label = "Todo el historial"

    body = f"""
<div class="hero-total">
  <div class="hero-kicker">PEDIDOS TALMA</div>
  <div class="hero-number">{len(rows)}</div>
  <div class="hero-date">menús pedidos · {periodo_label}</div>
</div>

<div class="grid grid3">
<div class="stat">Menús pedidos<b>{len(rows)}</b></div>
<div class="stat">Personas<b>{len(counts)}</b></div>
<div class="stat">Período<b>{periodo_label}</b></div>
</div>

<div class="card no-print">
<h2>📅 Filtrar por fecha</h2>
<form method="get" action="/talma" class="grid grid3">
  <div>
    <label>Desde</label>
    <input type="date" name="desde" value="{esc(fecha_desde)}">
  </div>
  <div>
    <label>Hasta</label>
    <input type="date" name="hasta" value="{esc(fecha_hasta)}">
  </div>
  <div style="align-self:end">
    <button type="submit">Ver período</button>
    <a class="btn secondary" href="/talma">Todo el historial</a>
  </div>
</form>
<p class="muted">Ejemplo: selecciona <b>8 de agosto</b> como inicio y <b>10 de agosto</b> como fin para saber cuántos menús se pidieron entre esas fechas.</p>
</div>

<div class="card no-print">
<h2>🔎 Buscar por DNI</h2>
<form method="get" action="/talma" class="actions">
  <input type="hidden" name="desde" value="{esc(fecha_desde)}">
  <input type="hidden" name="hasta" value="{esc(fecha_hasta)}">
  <input name="dni" value="{esc(dni_filter)}" inputmode="numeric" pattern="[0-9]{{8}}" maxlength="8" placeholder="DNI de 8 dígitos" style="max-width:260px">
  <button type="submit">Filtrar</button>
  <a class="btn secondary" href="/talma?desde={esc(fecha_desde)}&hasta={esc(fecha_hasta)}">Limpiar DNI</a>
</form>
<p class="muted">Para TALMA, el DNI es el identificador principal. Los registros antiguos sin DNI se muestran como "Sin DNI".</p>
</div>

<div class="card">
<h2>Resumen por DNI — {periodo_label}</h2>
<div class="table-wrap"><table><thead><tr><th>DNI</th><th>Persona</th><th>Total almuerzos</th></tr></thead>
<tbody>{''.join(summary) or '<tr><td colspan="3">No hay pedidos en este período.</td></tr>'}</tbody></table></div>
</div>

<div class="card">
<h2>Pedidos TALMA — {periodo_label}</h2>
<div class="table-wrap"><table><thead><tr><th>Fecha</th><th>DNI</th><th>Persona</th><th>Área</th><th>Entrada</th><th>Plato de fondo</th><th>N° Almuerzo</th><th>Hora</th></tr></thead>
<tbody>{''.join(table_rows) or '<tr><td colspan="8">No hay pedidos en este período.</td></tr>'}</tbody></table></div>
</div>
"""
    extra = """<style>
.hero-total{background:linear-gradient(135deg,#0f5132,#198754);color:white;border-radius:24px;padding:28px;text-align:center;margin-bottom:22px;box-shadow:0 14px 40px rgba(15,81,50,.20)}
.hero-kicker{font-weight:800;letter-spacing:3px;font-size:14px}.hero-number{font-size:68px;font-weight:900;line-height:1;margin:8px 0}.hero-date{font-size:18px;font-weight:700}
</style>"""
    return _talma_page("TALMA", body, extra)


@app.post("/talma/login")
def talma_login():
    password = request.form.get("password", "")
    if not hmac.compare_digest(password, _talma_password()):
        return _talma_login(True)
    value = _talma_session_value()
    secure = "; Secure" if request.headers.get("X-Forwarded-Proto") == "https" else ""
    response = redirect("/talma/")
    response.headers["Set-Cookie"] = f"talma_session={value}; Path=/; Max-Age=43200; HttpOnly; SameSite=Lax{secure}"
    return response


@app.get("/talma/logout")
def talma_logout():
    response = redirect("/talma")
    response.headers["Set-Cookie"] = "talma_session=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax"
    return response


@app.get("/talma/excel")
def talma_excel():
    if not _valid_talma_session(request.cookies.get("talma_session")):
        return _talma_login()
    orders_app.init_db()
    fecha_desde = request.args.get("desde", "").strip()
    fecha_hasta = request.args.get("hasta", "").strip()
    dni_filter = re.sub(r"\D", "", request.args.get("dni", ""))[:8]

    def _valid_iso(value):
        try:
            datetime.strptime(value, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    if fecha_desde and not _valid_iso(fecha_desde): fecha_desde = ""
    if fecha_hasta and not _valid_iso(fecha_hasta): fecha_hasta = ""
    if fecha_desde and not fecha_hasta: fecha_hasta = fecha_desde
    if fecha_hasta and not fecha_desde: fecha_desde = fecha_hasta
    if fecha_desde and fecha_hasta and fecha_desde > fecha_hasta:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    with orders_app.db() as conn:
        sql = """SELECT o.order_date, o.employee_name, o.dni, o.area, o.entry_item,
                      o.main_item, o.notes, o.created_at
               FROM orders o JOIN companies c ON c.id=o.company_id
               WHERE c.slug='talma'"""
        args = []
        if dni_filter:
            sql += " AND o.dni=?"; args.append(dni_filter)
        if fecha_desde:
            sql += " AND o.order_date>=?"; args.append(fecha_desde)
        if fecha_hasta:
            sql += " AND o.order_date<=?"; args.append(fecha_hasta)
        sql += " ORDER BY o.order_date, CASE WHEN o.dni='' THEN 1 ELSE 0 END, o.dni, o.employee_name COLLATE NOCASE, o.id"
        rows = conn.execute(sql, args).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos TALMA"
    headers = ["Fecha", "DNI", "Persona", "Área", "Entrada", "Plato de fondo", "Observación", "N° Almuerzo", "Hora"]
    ws.append(headers)
    counts = {}
    for r in rows:
        key = r["dni"] or f"legacy:{orders_app.normalize_key(r['employee_name'])}"
        counts[key] = counts.get(key, 0) + 1
        ws.append([r["order_date"], r["dni"], r["employee_name"], r["area"], r["entry_item"], r["main_item"],
                   r["notes"], counts[key], (r["created_at"] or "")[11:16]])
    for i, width in enumerate([14, 14, 34, 16, 32, 36, 40, 14, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="176B43")
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary = wb.create_sheet("Resumen")
    summary.append(["DNI", "PERSONA", "TOTAL ALMUERZOS"])
    for key, total in sorted(counts.items()):
        rr = next((r for r in rows if (r["dni"] or f"legacy:{orders_app.normalize_key(r['employee_name'])}") == key), None)
        summary.append([rr["dni"] if rr and rr["dni"] else "Sin DNI", rr["employee_name"] if rr else key, total])
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 34
    summary.column_dimensions["C"].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    suffix = f"_{fecha_desde}_a_{fecha_hasta}" if fecha_desde and fecha_hasta else "_historial_completo"
    return send_file(output, as_attachment=True, download_name=f"pedidos_talma{suffix}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Integración con el sistema principal de pedidos
# ---------------------------------------------------------------------------
_INTERNAL_ORDERS_PORT = int(os.environ.get("INTERNAL_ORDERS_PORT", "18081"))
_orders_server = None
_orders_thread = None
_orders_lock = threading.Lock()


def _ensure_orders_server():
    """Inicia el servidor original de pedidos en localhost una sola vez."""
    global _orders_server, _orders_thread
    with _orders_lock:
        if _orders_thread and _orders_thread.is_alive():
            return
        orders_app.init_db()
        _orders_server = ThreadingHTTPServer(("127.0.0.1", _INTERNAL_ORDERS_PORT), orders_app.AppHandler)
        _orders_thread = threading.Thread(target=_orders_server.serve_forever, name="pedidos-interno", daemon=True)
        _orders_thread.start()


@app.before_request
def _protect_scanner_admin():
    protected = (
        "/admin/talma",
        "/admin/policia",
        "/admin/scanners",
        "/admin/scanner-imagen",
        "/admin/scanner-diagnostico",
    )
    if request.path.startswith(protected):
        if not orders_app.valid_session(request.cookies.get("admin_session")):
            return redirect("/admin")


def _proxy_to_orders(path=""):
    """Reenvía al sistema de pedidos original cualquier ruta que no sea del scanner."""
    _ensure_orders_server()
    target = "/" + path if path else "/"
    if request.query_string:
        target += "?" + request.query_string.decode("latin-1")

    body = request.get_data(cache=False)
    headers = {}
    skipped = {"connection", "content-length", "transfer-encoding"}
    for key, value in request.headers.items():
        if key.lower() not in skipped:
            headers[key] = value
    # Mantener el host público para que los enlaces privados se generen con el dominio real.
    headers["Host"] = request.host
    headers["X-Forwarded-Proto"] = request.headers.get("X-Forwarded-Proto", request.scheme)

    connection = http.client.HTTPConnection("127.0.0.1", _INTERNAL_ORDERS_PORT, timeout=180)
    try:
        connection.request(request.method, target, body=body if body else None, headers=headers)
        upstream = connection.getresponse()
        data = upstream.read()
        response_headers = []
        for key, value in upstream.getheaders():
            if key.lower() not in {"connection", "transfer-encoding", "content-length", "server", "date"}:
                response_headers.append((key, value))
        return Response(data, status=upstream.status, headers=response_headers)
    finally:
        connection.close()


@app.route("/", defaults={"path": ""}, methods=["GET", "POST", "HEAD"])
@app.route("/<path:path>", methods=["GET", "POST", "HEAD"])
def orders_proxy(path):
    return _proxy_to_orders(path)


# Se inicia también al importar la aplicación (por ejemplo, con Gunicorn).
_ensure_orders_server()


if __name__ == "__main__":
    public_port = int(os.environ.get("PORT", "8080"))
    app.run(host="0.0.0.0", port=public_port, debug=False, threaded=True)
