#!/usr/bin/env python3
"""
Sistema sencillo de pedidos de almuerzos para empresas.
No requiere librerías externas: usa solo Python + SQLite.
"""
from __future__ import annotations

from datetime import datetime, date

import csv
from collections import Counter
import hashlib
import hmac
import html
import io
import json
import os
import re
import secrets
import sqlite3
import sys
import threading
import time
import unicodedata
from datetime import date, datetime
from zoneinfo import ZoneInfo
from http import cookies
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlencode, urlparse
from email.parser import BytesParser
from email.policy import default as email_default_policy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR / "datos"))).expanduser().resolve()
DB_PATH = DATA_DIR / "pedidos.db"
CONFIG_PATH = BASE_DIR / "config.json"
HOST = "0.0.0.0"
try:
    PORT = int(os.environ.get("PORT", "8080"))
except ValueError:
    PORT = 8080

DATA_DIR.mkdir(exist_ok=True)

DEFAULT_CONFIG = {
    "restaurant_name": "Mi Restaurante",
    "admin_password": "cambiar123",
    "secret_key": secrets.token_hex(32),
    "order_deadline": "11:30",
}

# Empresas que siempre deben existir en el sistema.
# El token privado se genera solo la primera vez y se conserva en la base de datos.
DEFAULT_COMPANIES = [
    ("Liderman", "liderman"),
    ("Aeropuerto", "aeropuerto"),
    ("Talma", "talma"),
    ("Policía", "policia"),
]


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        CONFIG_PATH.write_text(json.dumps(DEFAULT_CONFIG, indent=2, ensure_ascii=False), encoding="utf-8")
    try:
        cfg = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        cfg = DEFAULT_CONFIG.copy()
    changed = False
    for key, value in DEFAULT_CONFIG.items():
        if key not in cfg:
            cfg[key] = value
            changed = True
    if changed:
        CONFIG_PATH.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")

    # En producción, las variables de entorno tienen prioridad sobre config.json.
    env_overrides = {
        "restaurant_name": "RESTAURANT_NAME",
        "admin_password": "ADMIN_PASSWORD",
        "secret_key": "SECRET_KEY",
        "order_deadline": "ORDER_DEADLINE",
    }
    for config_key, env_key in env_overrides.items():
        env_value = os.environ.get(env_key)
        if env_value:
            cfg[config_key] = env_value
    return cfg


CONFIG = load_config()

TALMA_ROSTER_PATH = BASE_DIR / "talma_personas.xlsx"

def load_talma_roster() -> dict:
    """Carga exclusivamente la base oficial de personas TALMA desde talma_personas.xlsx."""
    roster = {}
    try:
        workbook = load_workbook(TALMA_ROSTER_PATH, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(v or "").strip().upper() for v in next(rows)]
        dni_idx = headers.index("DNI")
        name_idx = headers.index("NOMBRE_APELLIDOS")
        area_idx = headers.index("AREA")
        email_idx = headers.index("EMAIL")
        for row in rows:
            if not row:
                continue
            raw_dni = row[dni_idx]
            dni = str(raw_dni).strip() if raw_dni is not None else ""
            if dni.endswith(".0"):
                dni = dni[:-2]
            name = str(row[name_idx] or "").strip()
            area = str(row[area_idx] or "").strip()
            email = str(row[email_idx] or "").strip().lower()
            if dni and email:
                roster[f"{email}|{dni}"] = {
                    "email": email, "dni": dni, "nombre": name, "area": area
                }
        workbook.close()
    except Exception as error:
        print(f"[TALMA] No se pudo leer la base Excel de personas: {error}", file=sys.stderr)
    return roster

TALMA_ROSTER = load_talma_roster()



def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def hash_talma_password(password: str, salt_hex: str | None = None) -> tuple[str, str]:
    salt = bytes.fromhex(salt_hex) if salt_hex else os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 180_000)
    return digest.hex(), salt.hex()


def verify_talma_password(password: str, password_hash: str, password_salt: str) -> bool:
    try:
        digest, _ = hash_talma_password(password, password_salt)
        return hmac.compare_digest(digest, password_hash)
    except Exception:
        return False


def get_talma_manual_user(email: str, dni: str, include_inactive: bool = False):
    return None

def init_db() -> None:
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                slug TEXT NOT NULL UNIQUE,
                token TEXT NOT NULL UNIQUE,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS menu_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                menu_date TEXT NOT NULL,
                category TEXT NOT NULL CHECK(category IN ('entrada','fondo')),
                name TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                UNIQUE(menu_date, category, name)
            );

            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                order_date TEXT NOT NULL,
                employee_name TEXT NOT NULL,
                employee_key TEXT NOT NULL,
                dni TEXT NOT NULL DEFAULT '',
                area TEXT NOT NULL DEFAULT '',
                entry_item TEXT NOT NULL,
                main_item TEXT NOT NULL,
                notes TEXT NOT NULL DEFAULT '',
                delivery_type TEXT NOT NULL DEFAULT 'Empresa',
                created_at TEXT NOT NULL,
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE,
                UNIQUE(company_id, order_date, employee_key)
            );
            """
        )

        # La antigua tabla de personas TALMA se elimina: la fuente oficial es el Excel.
        conn.execute("DROP TABLE IF EXISTS talma_manual_users")

        # Migración compatible: versiones anteriores no tenían DNI.
        columns = {row["name"] for row in conn.execute("PRAGMA table_info(orders)").fetchall()}
        if "dni" not in columns:
            conn.execute("ALTER TABLE orders ADD COLUMN dni TEXT NOT NULL DEFAULT ''")
        if "login_code" not in columns:
            conn.execute("ALTER TABLE orders ADD COLUMN login_code TEXT NOT NULL DEFAULT ''")
        if "login_email" not in columns:
            conn.execute("ALTER TABLE orders ADD COLUMN login_email TEXT NOT NULL DEFAULT ''")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_orders_talma_dni ON orders(dni)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_orders_talma_code ON orders(login_code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_orders_talma_email ON orders(login_email)")
        # Elimina el enlace de demostración de versiones anteriores y garantiza
        # los cuatro enlaces predeterminados solicitados.
        conn.execute("DELETE FROM companies WHERE slug='empresa-demo'")
        for company_name, company_slug in DEFAULT_COMPANIES:
            existing = conn.execute(
                "SELECT id FROM companies WHERE slug=?",
                (company_slug,),
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE companies SET name=?, active=1 WHERE id=?",
                    (company_name, existing["id"]),
                )
            else:
                conn.execute(
                    "INSERT INTO companies(name, slug, token, created_at) VALUES (?, ?, ?, ?)",
                    (company_name, company_slug, secrets.token_urlsafe(18), now_iso()),
                )

        today = today_iso()
        menu_count = conn.execute("SELECT COUNT(*) FROM menu_items WHERE menu_date=?", (today,)).fetchone()[0]
        if menu_count == 0:
            demo = [
                (today, "entrada", "Sopa del día"),
                (today, "entrada", "Ensalada fresca"),
                (today, "fondo", "Pollo al horno con arroz"),
                (today, "fondo", "Lomo saltado"),
            ]
            conn.executemany("INSERT INTO menu_items(menu_date, category, name) VALUES (?, ?, ?)", demo)

    start_backup_worker()


LOCAL_TZ = ZoneInfo("America/Lima")


def local_now() -> datetime:
    return datetime.now(LOCAL_TZ)


def today_iso() -> str:
    return local_now().date().isoformat()


def now_iso() -> str:
    return local_now().replace(tzinfo=None, microsecond=0).isoformat(sep=" ")


# ---------------------------------------------------------------------------
# Persistencia y copias de seguridad
# ---------------------------------------------------------------------------
BACKUP_DIR = DATA_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

def backup_database() -> Path | None:
    """Crea una copia consistente diaria de SQLite y conserva 30 días."""
    try:
        target = BACKUP_DIR / f"pedidos_{today_iso()}.db"
        if target.exists():
            return target
        temp = BACKUP_DIR / f".pedidos_{today_iso()}.tmp"
        if temp.exists():
            temp.unlink()
        with db() as source:
            with sqlite3.connect(temp) as destination:
                source.backup(destination)
        os.replace(temp, target)
        backups = sorted(BACKUP_DIR.glob("pedidos_*.db"), key=lambda x: x.name, reverse=True)
        for old in backups[30:]:
            try:
                old.unlink()
            except OSError:
                pass
        return target
    except Exception as error:
        print(f"[BACKUP] No se pudo crear la copia: {error}", file=sys.stderr)
        return None

def _backup_worker() -> None:
    while True:
        try:
            backup_database()
            time.sleep(3600)
        except Exception as error:
            print(f"[BACKUP] Error en el proceso diario: {error}", file=sys.stderr)
            time.sleep(300)

_backup_thread_started = False
_backup_thread_lock = threading.Lock()

def start_backup_worker() -> None:
    global _backup_thread_started
    with _backup_thread_lock:
        if _backup_thread_started:
            return
        _backup_thread_started = True
        backup_database()
        threading.Thread(target=_backup_worker, name="backup-diario", daemon=True).start()


def is_order_closed(order_date: str) -> bool:
    """Cierra los pedidos del día actual a las 11:30, hora de Perú."""
    if order_date != today_iso():
        return False
    now = local_now()
    return (now.hour, now.minute, now.second) >= (11, 30, 0)

def seconds_until_close() -> int:
    now = local_now()
    close = now.replace(hour=11, minute=30, second=0, microsecond=0)
    return max(0, int((close-now).total_seconds()))


def normalize_key(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def slugify(text: str) -> str:
    slug = normalize_key(text).replace(" ", "-")
    return slug or f"empresa-{secrets.token_hex(3)}"


def esc(value: object) -> str:
    return html.escape(str(value), quote=True)


def page(title: str, body: str, extra_head: str = "") -> str:
    restaurant = esc(CONFIG.get("restaurant_name", "Mi Restaurante"))
    return f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(title)} · {restaurant}</title>
<style>
:root{{--bg:#f4f6f8;--card:#fff;--text:#18212b;--muted:#667085;--brand:#175cd3;--brand2:#0b4bab;--danger:#b42318;--ok:#067647;--line:#e4e7ec}}
*{{box-sizing:border-box}} body{{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:var(--bg);color:var(--text)}}
a{{color:var(--brand);text-decoration:none}} .topbar{{background:#101828;color:#fff;padding:14px 20px;display:flex;justify-content:space-between;align-items:center;gap:15px}}
.topbar strong{{font-size:18px}} .topbar a{{color:#fff}} .wrap{{max-width:1180px;margin:24px auto;padding:0 16px}} .narrow{{max-width:720px}}
.card{{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:20px;margin-bottom:18px;box-shadow:0 2px 10px rgba(16,24,40,.04)}}
h1,h2,h3{{margin-top:0}} h1{{font-size:28px}} h2{{font-size:21px}} .muted{{color:var(--muted)}} .grid{{display:grid;gap:16px}} .grid2{{grid-template-columns:repeat(2,minmax(0,1fr))}} .grid3{{grid-template-columns:repeat(3,minmax(0,1fr))}}
label{{display:block;font-weight:600;margin:12px 0 6px}} input,select,textarea{{width:100%;padding:11px 12px;border:1px solid #d0d5dd;border-radius:9px;font:inherit;background:#fff}} textarea{{min-height:96px;resize:vertical}}
button,.btn{{display:inline-block;border:0;border-radius:9px;padding:11px 15px;background:var(--brand);color:#fff;font-weight:700;cursor:pointer}} button:hover,.btn:hover{{background:var(--brand2)}} .btn.secondary{{background:#475467}} .btn.danger{{background:var(--danger)}} .btn.small{{padding:7px 10px;font-size:13px}}
.notice{{padding:12px 14px;border-radius:9px;margin:12px 0}} .notice.ok{{background:#ecfdf3;color:var(--ok);border:1px solid #abefc6}} .notice.error{{background:#fef3f2;color:var(--danger);border:1px solid #fecdca}}
table{{width:100%;border-collapse:collapse;font-size:14px}} th,td{{border-bottom:1px solid var(--line);padding:10px 8px;text-align:left;vertical-align:top}} th{{background:#f9fafb}} .table-wrap{{overflow:auto}} .actions{{display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
.stat{{padding:16px;border:1px solid var(--line);border-radius:12px;background:#fff}} .stat b{{display:block;font-size:26px;margin-top:5px}}
.menu-choice{{border:1px solid var(--line);border-radius:10px;padding:12px;margin:8px 0}} .menu-choice input{{width:auto;margin-right:8px}} .ticket{{border:2px dashed #333;padding:12px;margin:0 0 12px;break-inside:avoid;background:#fff}} .ticket h3{{margin-bottom:8px}}
.registered-orders{{margin:0;padding-left:22px}} .registered-orders li{{padding:8px 0;border-bottom:1px solid var(--line)}} .registered-orders li:last-child{{border-bottom:0}}
footer{{text-align:center;color:var(--muted);padding:28px 16px;font-size:13px;line-height:1.6;border-top:1px solid var(--line);margin-top:28px}} .support-footer{{max-width:1180px;margin:0 auto}}
.admin-tabs{{display:flex;gap:8px;flex-wrap:wrap;background:#fff;border:1px solid var(--line);border-radius:12px;padding:8px;margin:0 0 18px;box-shadow:0 2px 10px rgba(16,24,40,.04)}}
.admin-tabs a{{padding:10px 16px;border-radius:9px;font-weight:800;color:#344054}} .admin-tabs a:hover{{background:#f2f4f7}} .admin-tabs a.active{{background:var(--brand);color:#fff}}
@media(max-width:760px){{.grid2,.grid3{{grid-template-columns:1fr}} .topbar{{align-items:flex-start;flex-direction:column}}}}
@media print{{.no-print,.topbar,footer{{display:none!important}} body{{background:#fff}} .wrap{{max-width:none;margin:0;padding:0}} .card{{border:0;box-shadow:none;padding:0}} .ticket{{page-break-inside:avoid}}}}
</style>
{extra_head}
</head>
<body>
<div class="topbar"><strong>{restaurant}</strong></div>
{body}
<footer>
<div class="support-footer">
<strong>Soporte del sistema</strong><br>
Cualquier duda o sugerencia respecto al sistema, comunicarse por:<br>
931099267 &nbsp;|&nbsp; 927314911 &nbsp;|&nbsp; sebasoa0711@gmail.com
<br><span>Todos los derechos reservados · Urpicha Restaurante</span>
</div>
</footer>
</body></html>"""


def get_company(slug: str, token: str | None = None):
    with db() as conn:
        if token is None:
            return conn.execute("SELECT * FROM companies WHERE slug=? AND active=1", (slug,)).fetchone()
        return conn.execute("SELECT * FROM companies WHERE slug=? AND token=? AND active=1", (slug, token)).fetchone()


def get_menu(menu_date: str):
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM menu_items WHERE menu_date=? AND active=1 ORDER BY category, id", (menu_date,)
        ).fetchall()
    return {
        "entrada": [r for r in rows if r["category"] == "entrada"],
        "fondo": [r for r in rows if r["category"] == "fondo"],
    }


def session_value() -> str:
    expires = str(int(time.time()) + 8 * 3600)
    payload = f"admin:{expires}"
    sig = hmac.new(CONFIG["secret_key"].encode(), payload.encode(), hashlib.sha256).hexdigest()
    return f"{payload}:{sig}"


def valid_session(value: str | None) -> bool:
    if not value:
        return False
    try:
        user, expires, sig = value.split(":", 2)
        if user != "admin" or int(expires) < time.time():
            return False
        payload = f"{user}:{expires}"
        expected = hmac.new(CONFIG["secret_key"].encode(), payload.encode(), hashlib.sha256).hexdigest()
        return hmac.compare_digest(sig, expected)
    except (ValueError, TypeError):
        return False





def resolve_talma_login(email: str, dni: str):
    email = email.strip().lower()
    dni = dni.strip()
    if not email or not dni:
        return None
    record = TALMA_ROSTER.get(f"{email}|{dni}")
    if record:
        result = dict(record)
        result["manual"] = False
        return result
    return None

def talma_employee_session_value(record: dict) -> str:
    expires = str(int(time.time()) + 12 * 3600)
    payload = f"talma_employee:{record['email']}:{record['dni']}:{expires}"
    sig = hmac.new(CONFIG["secret_key"].encode(), payload.encode(), hashlib.sha256).hexdigest()
    return f"{payload}:{sig}"


def valid_talma_employee_session(value: str | None) -> dict | None:
    if not value:
        return None
    try:
        user, email, dni, expires, sig = value.split(":", 4)
        if user != "talma_employee" or int(expires) < time.time():
            return None
        payload = f"{user}:{email}:{dni}:{expires}"
        expected = hmac.new(CONFIG["secret_key"].encode(), payload.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sig, expected):
            return None
        return resolve_talma_login(email, dni)
    except (ValueError, TypeError):
        return None


def talma_employee_login_page(token: str, error: str = "") -> str:
    notice = f'<div class="notice error">{esc(error)}</div>' if error else ""
    body = f"""
<main class="talma-login-shell">
<div class="talma-login-card">
<div class="talma-brand">TALMA</div>
<div class="talma-line"></div>
<h1>Acceso a pedidos</h1>
<p class="talma-subtitle">Ingresa con tu correo y DNI registrados.</p>
{notice}
<form method="post" action="/empresa/talma/login">
<input type="hidden" name="token" value="{esc(token)}">
<label>Correo electrónico</label>
<input type="email" name="email" autocomplete="username" required placeholder="correo@empresa.com">
<label>DNI</label>
<input type="text" name="dni" inputmode="numeric" autocomplete="username" required placeholder="DNI" maxlength="8">
<p class="talma-help">El sistema valida ambos datos y carga automáticamente tu nombre y área.</p>
<button type="submit">Ingresar</button>
</form>
</div>
</main>"""
    extra="""<style>
.talma-login-shell{min-height:100vh;display:flex;align-items:center;justify-content:center;padding:30px;background:linear-gradient(145deg,#f4f8fb 0%,#eef6f7 100%)}
.talma-login-card{width:min(440px,100%);background:#fff;border:1px solid #d5e1e5;border-radius:18px;padding:34px;box-shadow:0 18px 50px rgba(19,60,70,.12)}
.talma-brand{font-size:28px;font-weight:900;letter-spacing:4px;color:#006b78}.talma-line{height:4px;width:62px;background:#00a7b5;border-radius:6px;margin:10px 0 24px}
.talma-login-card h1{font-size:25px;margin-bottom:8px;color:#18343b}.talma-subtitle{color:#667085;margin-top:0;margin-bottom:22px}
.talma-login-card input{margin-bottom:6px}.talma-login-card button{width:100%;margin-top:12px;background:#006b78}.talma-login-card button:hover{background:#005761}
.talma-help{font-size:12px;color:#7b8794;margin-top:10px;line-height:1.45}
</style>"""
    return page("Acceso TALMA", body, extra)

def normalize_report_header(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def read_multipart_file(handler: BaseHTTPRequestHandler, field_name: str, max_bytes: int = 20_000_000):
    """Lee un único archivo de un formulario multipart sin depender del módulo cgi."""
    try:
        length = int(handler.headers.get("Content-Length", "0"))
    except ValueError:
        length = 0
    if length <= 0 or length > max_bytes:
        raise ValueError("El archivo está vacío o supera el límite de 20 MB.")
    raw = handler.rfile.read(length)
    content_type = handler.headers.get("Content-Type", "")
    if "multipart/form-data" not in content_type:
        raise ValueError("El formulario de archivo no es multipart/form-data.")
    mime = b"Content-Type: " + content_type.encode("utf-8") + b"\r\nMIME-Version: 1.0\r\n\r\n" + raw
    message = BytesParser(policy=email_default_policy).parsebytes(mime)
    for part in message.iter_parts():
        if part.get_content_disposition() == "form-data" and part.get_param("name", header="content-disposition") == field_name:
            filename = part.get_filename() or "historico.xlsx"
            return filename, part.get_payload(decode=True) or b""
    raise ValueError("No se encontró el archivo Excel.")


def extract_historical_orders(workbook):
    """Extrae filas de un Excel anterior usando encabezados flexibles."""
    aliases = {
        "fecha": {"fecha", "date", "dia", "fecha pedido"},
        "empresa": {"empresa", "company", "compania"},
        "nombre": {"empleado", "nombre", "nombre y apellido", "persona", "trabajador"},
        "area": {"area", "area o sede", "sede"},
        "entrada": {"entrada", "plato de entrada"},
        "fondo": {"plato de fondo", "fondo", "segundo", "plato fondo", "main"},
        "observacion": {"observacion", "observaciones", "sugerencias", "notas"},
    }
    records = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_idx = None
        mapping = {}
        for i, raw_header in enumerate(rows[:10]):
            normalized = [normalize_report_header(v) for v in raw_header]
            candidate = {}
            for idx, h in enumerate(normalized):
                for field, names in aliases.items():
                    if h in names and field not in candidate:
                        candidate[field] = idx
            if "nombre" in candidate and ("fecha" in candidate or "entrada" in candidate):
                header_idx = i
                mapping = candidate
                break
        if header_idx is None:
            continue

        for raw in rows[header_idx + 1:]:
            def val(field):
                idx = mapping.get(field)
                return raw[idx] if idx is not None and idx < len(raw) else ""

            name = re.sub(r"\s+", " ", str(val("nombre") or "")).strip()
            entry = re.sub(r"\s+", " ", str(val("entrada") or "")).strip()
            fondo = re.sub(r"\s+", " ", str(val("fondo") or "")).strip()
            if not name and not entry and not fondo:
                continue

            raw_date = val("fecha")
            if isinstance(raw_date, datetime):
                order_date = raw_date.strftime("%Y-%m-%d")
            elif isinstance(raw_date, date):
                order_date = raw_date.isoformat()
            else:
                parsed = None
                raw_text = re.sub(r"\s+", " ", str(raw_date or "")).strip()
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
                    try:
                        parsed = datetime.strptime(raw_text, fmt)
                        break
                    except ValueError:
                        pass
                order_date = parsed.strftime("%Y-%m-%d") if parsed else raw_text

            records.append({
                "fecha": order_date,
                "empresa": re.sub(r"\s+", " ", str(val("empresa") or "")).strip(),
                "nombre": name,
                "area": re.sub(r"\s+", " ", str(val("area") or "")).strip(),
                "entrada": entry,
                "fondo": fondo,
                "observacion": re.sub(r"\s+", " ", str(val("observacion") or "")).strip(),
            })
    return records


def build_historical_report(workbook, historical_records, current_records):
    """Añade al Excel un detalle unificado y reportes por persona y por día."""
    all_records = []
    seen = set()

    for record in historical_records + current_records:
        key = (
            normalize_key(record.get("fecha", "")),
            normalize_key(record.get("empresa", "")),
            normalize_key(record.get("nombre", "")),
            normalize_key(record.get("entrada", "")),
            normalize_key(record.get("fondo", "")),
        )
        if not key[0] or not key[2]:
            continue
        if key in seen:
            continue
        seen.add(key)
        all_records.append(record)

    # El detalle queda ordenado alfabéticamente por persona para identificarla rápidamente.
    # Dentro de cada persona, sus almuerzos se ordenan por fecha y empresa para que
    # el N° Almuerzo siga una secuencia lógica.
    all_records.sort(key=lambda r: (normalize_key(r.get("nombre", "")), normalize_key(r.get("empresa", "")), r.get("fecha", "")))

    # Reemplaza las hojas de reporte para que la actualización sea idempotente.
    for name in ("REPORTE PEDIDOS", "DETALLE UNIFICADO"):
        if name in workbook.sheetnames:
            del workbook[name]

    detail = workbook.create_sheet("DETALLE UNIFICADO")
    headers = ["Fecha", "Empresa", "Persona", "Área", "Entrada", "Plato de fondo", "Observación", "N° Almuerzo"]
    detail.append(headers)

    person_counts = {}
    for record in all_records:
        person_key = (normalize_key(record["empresa"]), normalize_key(record["nombre"]))
        person_counts[person_key] = person_counts.get(person_key, 0) + 1

    running = {}
    for record in all_records:
        person_key = (normalize_key(record["empresa"]), normalize_key(record["nombre"]))
        running[person_key] = running.get(person_key, 0) + 1
        detail.append([
            record["fecha"], record["empresa"], record["nombre"], record["area"],
            record["entrada"], record["fondo"], record["observacion"], running[person_key]
        ])

    report = workbook.create_sheet("REPORTE PEDIDOS")
    report.append(["REPORTE HISTÓRICO DE ALMUERZOS"])
    report.append([f"Generado: {local_now().strftime('%d/%m/%Y %H:%M')}"])
    report.append([f"Total de registros: {len(all_records)}"])
    report.append([])
    report.append(["RESUMEN POR PERSONA"])
    report.append(["Empresa", "Persona", "Total almuerzos", "Último pedido"])

    by_person = {}
    for record in all_records:
        key = (record["empresa"], record["nombre"])
        item = by_person.setdefault(key, {"total": 0, "last": "", "entries": Counter(), "mains": Counter()})
        item["total"] += 1
        item["last"] = max(item["last"], record["fecha"])
        if record["entrada"]:
            item["entries"][record["entrada"]] += 1
        if record["fondo"]:
            item["mains"][record["fondo"]] += 1

    for (empresa, nombre), item in sorted(by_person.items(), key=lambda x: (normalize_key(x[0][0]), normalize_key(x[0][1]))):
        report.append([empresa, nombre, item["total"], item["last"]])

    report.append([])
    report.append(["RESUMEN POR DÍA"])
    report.append(["Fecha", "Empresa", "Pedidos"])
    by_day = Counter((r["fecha"], r["empresa"]) for r in all_records)
    for (fecha, empresa), total in sorted(by_day.items()):
        report.append([fecha, empresa, total])

    report.append([])
    report.append(["PLATOS MÁS PEDIDOS"])
    report.append(["Tipo", "Plato", "Cantidad"])
    dish_counts = Counter()
    for record in all_records:
        if record["entrada"]:
            dish_counts[("Entrada", record["entrada"])] += 1
        if record["fondo"]:
            dish_counts[("Plato de fondo", record["fondo"])] += 1
    for (kind, dish), total in dish_counts.most_common():
        report.append([kind, dish, total])

    # Formato
    header_fill = PatternFill("solid", fgColor="176B43")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D5DD")
    for sheet in (detail, report):
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
    if detail.max_row > 1:
        detail.auto_filter.ref = detail.dimensions
    if report.max_row > 1:
        # El resumen por persona también queda alfabético.
        report.auto_filter.ref = f"A6:D{5 + len(by_person) + 1}" if by_person else "A6:D6"

    for idx, width in enumerate([14, 18, 28, 18, 32, 36, 30, 14], 1):
        detail.column_dimensions[get_column_letter(idx)].width = width
    for idx, width in enumerate([20, 30, 18, 18], 1):
        report.column_dimensions[get_column_letter(idx)].width = width

    return len(all_records), len(by_person)


class AppHandler(BaseHTTPRequestHandler):
    server_version = "Almuerzos/1.0"

    def log_message(self, fmt: str, *args) -> None:
        sys.stdout.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))

    def parse_cookies(self) -> cookies.SimpleCookie:
        jar = cookies.SimpleCookie()
        if self.headers.get("Cookie"):
            jar.load(self.headers["Cookie"])
        return jar

    def is_admin(self) -> bool:
        jar = self.parse_cookies()
        value = jar.get("admin_session")
        return valid_session(value.value if value else None)

    def read_form(self) -> dict[str, str]:
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            length = 0
        raw = self.rfile.read(min(length, 2_000_000)).decode("utf-8", "replace")
        parsed = parse_qs(raw, keep_blank_values=True)
        return {k: v[0] if v else "" for k, v in parsed.items()}

    def send_html(self, content: str, status: int = 200, headers: dict | None = None) -> None:
        data = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "SAMEORIGIN")
        self.send_header("Referrer-Policy", "same-origin")
        if headers:
            for key, value in headers.items():
                self.send_header(key, value)
        self.end_headers()
        self.wfile.write(data)

    def send_text(self, text: str, status: int = 200) -> None:
        data = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_bytes(self, data: bytes, content_type: str, filename: str) -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def redirect(self, location: str, cookie_header: str | None = None) -> None:
        self.send_response(303)
        self.send_header("Location", location)
        if cookie_header:
            self.send_header("Set-Cookie", cookie_header)
        self.end_headers()

    def require_admin(self) -> bool:
        if not self.is_admin():
            self.redirect("/admin")
            return False
        return True

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path.rstrip("/") or "/"
        query = {k: v[0] for k, v in parse_qs(parsed.query).items()}

        if path == "/":
            self.home()
        elif path == "/health":
            self.send_text("ok")
        elif path == "/admin":
            self.admin_login(query.get("error"))
        elif path == "/admin/logout":
            self.redirect("/admin", "admin_session=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax")
        elif path == "/admin/dashboard":
            if self.require_admin():
                self.admin_dashboard(query)
        elif path == "/admin/personas":
            if self.require_admin():
                self.admin_personas(query)
        elif path == "/admin/resumen":
            if self.require_admin():
                self.admin_resumen(query)
        elif path == "/admin/export.csv":
            if self.require_admin():
                self.export_csv(query)
        elif path == "/admin/cupones":
            if self.require_admin():
                self.coupons(query)
        elif path.startswith("/empresa/") and path.endswith("/excel"):
            self.company_export_excel(path.split("/")[2], query)
        elif path == "/empresa/talma/login":
            self.send_html(talma_employee_login_page(query.get("token", ""), query.get("error", "")))
        elif path == "/empresa/talma/logout":
            self.redirect(f"/empresa/talma/login?{urlencode({'token':query.get('token','')})}", "talma_employee_session=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax")
        elif path.startswith("/empresa/") and path.endswith("/pdf"):
            self.company_export_pdf(path.split("/")[2], query)
        elif path.startswith("/empresa/"):
            self.employee_form(path.split("/", 2)[2], query)
        else:
            self.send_html(page("No encontrado", '<main class="wrap narrow"><div class="card"><h1>404</h1><p>Página no encontrada.</p></div></main>'), 404)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path.rstrip("/") or "/"
        if path == "/admin/login":
            self.login()
        elif path == "/admin/menu":
            if self.require_admin():
                self.save_menu()
        elif path == "/admin/empresa":
            if self.require_admin():
                self.create_company()
        elif path == "/admin/empresa/toggle":
            if self.require_admin():
                self.toggle_company()
        elif path == "/admin/pedido/eliminar":
            if self.require_admin():
                self.delete_order()
        elif path == "/admin/historico-excel":
            if self.require_admin():
                self.update_historical_excel()
        elif path == "/empresa/talma/login":
            self.talma_employee_login()
        elif path.startswith("/pedido/"):
            self.submit_order(path.split("/", 2)[2])
        else:
            self.send_html(page("No encontrado", '<main class="wrap narrow"><div class="card"><h1>404</h1></div></main>'), 404)

    def home(self) -> None:

        body = f"""
<main class="wrap narrow">
<div class="card">
<h1>Pedidos de almuerzo</h1>
<p>Los trabajadores deben ingresar mediante el enlace privado de su empresa.</p>
<p class="muted">Administrador del restaurante: use el panel para publicar el menú y revisar pedidos.</p>
<a class="btn" href="/admin">Ingresar al panel administrador</a>
</div>
</main>"""
        self.send_html(page("Inicio", body))

    def whatsapp_link(self, message: str = "") -> str:
        number=re.sub(r"\D","",os.environ.get("WHATSAPP_NUMBER",str(CONFIG.get("whatsapp_number",""))))
        return f"https://wa.me/{number}?text={quote(message)}" if number else ""

    def talma_employee_login(self) -> None:
        form = self.read_form()
        token = form.get("token", "").strip()
        email = form.get("email", "").strip().lower()
        dni = form.get("dni", "").strip()
        record = resolve_talma_login(email, dni)
        company = get_company("talma", token) if token else get_company("talma")
        if company:
            token = company["token"]
        if not company:
            self.send_html(page("Enlace inválido", '<main class="wrap narrow"><div class="card"><h1>Enlace inválido</h1><p>El enlace de TALMA no es válido.</p></div></main>'), 403)
            return
        if not record:
            self.redirect(f"/empresa/talma/login?{urlencode({'token':token,'error':'Correo o DNI incorrecto.'})}")
            return
        value = talma_employee_session_value(record)
        secure = "; Secure" if self.headers.get("X-Forwarded-Proto") == "https" else ""
        location = f"/empresa/talma?{urlencode({'token':token,'ok_login':'1'})}"
        self.redirect(location, f"talma_employee_session={value}; Path=/; Max-Age=43200; HttpOnly; SameSite=Lax{secure}")

    def display_area(self, company_slug: str, area: str) -> str:
        value = str(area or "").strip()
        if company_slug.lower() == "talma" and value.upper() == "PAXHANDLING":
            return "PAX"
        return value

    def employee_form(self, slug: str, query: dict[str, str]) -> None:
        token = query.get("token", "").strip()
        company = get_company(slug, token) if token else get_company(slug)
        if not company:
            self.send_html(page("Enlace inválido", '<main class="wrap narrow"><div class="card"><h1>Enlace inválido</h1><p>Solicite a su empresa un enlace actualizado.</p></div></main>'), 403)
            return
        token = company["token"]
        is_talma = slug.lower() == "talma"
        talma_employee = None
        if is_talma:
            talma_employee = valid_talma_employee_session(self.parse_cookies().get("talma_employee_session").value if self.parse_cookies().get("talma_employee_session") else None)
            if not talma_employee:
                self.redirect(f"/empresa/talma/login?{urlencode({'token':token})}")
                return
        requested_date = query.get("fecha", today_iso())
        try:
            datetime.strptime(requested_date, "%Y-%m-%d")
        except ValueError:
            requested_date = today_iso()
        menu = get_menu(requested_date)
        with db() as conn:
            public_orders = conn.execute(
                """SELECT employee_name, dni, created_at FROM orders
                   WHERE company_id=? AND order_date=?
                   ORDER BY created_at DESC, id DESC""",
                (company["id"], requested_date),
            ).fetchall()
        notice = ""
        if query.get("ok") == "1":
            notice = '<div class="notice ok">Pedido registrado correctamente.</div>'
        elif query.get("error"):
            notice = f'<div class="notice error">{esc(query["error"])}</div>'

        if is_order_closed(requested_date):
            wa=self.whatsapp_link(f"Hola, necesito realizar un pedido fuera de horario para {company['name']}.")
            wa_button=(f'<a class="btn" target="_blank" rel="noopener" href="{esc(wa)}"> Comunicarme por WhatsApp</a>'
                       if wa else '<p><b>Comunícate por WhatsApp con el responsable de la empresa para coordinar la entrega.</b></p>')
            menu_html=f"""<div class="notice error">
<b> Pedidos cerrados</b>
<p>Los pedidos para <b>{esc(fecha_con_dia(requested_date))}</b> se cerraron a las <b>11:30 a. m.</b> (hora de Perú).</p>
{wa_button}
</div>"""
            submit=""
        elif not menu["entrada"] or not menu["fondo"]:
            menu_html = '<div class="notice error">Todavía no se ha publicado un menú completo para esta fecha.</div>'
            submit = ""
        else:
            entries = "".join(
                f'<label class="menu-choice"><input required type="radio" name="entrada" value="{esc(item["name"])}"> {esc(item["name"])}</label>'
                for item in menu["entrada"]
            )
            mains = "".join(
                f'<label class="menu-choice"><input required type="radio" name="fondo" value="{esc(item["name"])}"> {esc(item["name"])}</label>'
                for item in menu["fondo"]
            )
            menu_html = f"""
<div class="grid grid2">
<div><h3>Entrada</h3>{entries}</div>
<div><h3>Plato de fondo</h3>{mains}</div>
</div>"""
            submit = '<button type="submit">Enviar mi pedido</button>'

        if public_orders:
            public_order_rows = "".join(
                f'<li><b>{esc(o["employee_name"])}</b> — pedido registrado a las <b>{esc(o["created_at"][11:16])}</b></li>'
                for o in public_orders
            )
            public_orders_html = f"""
<div class="card">
<h2>Pedidos registrados</h2>
<p class="muted">Registros de {esc(company['name'])} para la fecha seleccionada.</p>
<ul class="registered-orders">{public_order_rows}</ul>
</div>"""
        else:
            public_orders_html = """
<div class="card">
<h2>Pedidos registrados</h2>
<p class="muted">Todavía no hay pedidos registrados para esta fecha.</p>
</div>"""

        if requested_date == today_iso() and not is_order_closed(requested_date):
            remaining = seconds_until_close()
            countdown_html = f"""<div class="countdown-box" style="background:#fff7ed;border:2px solid #f79009;border-radius:14px;padding:14px;text-align:center;margin:14px 0;color:#7a2e0e">
<b> Quedan <span id="countdown"></span> para el cierre de pedidos de hoy a las 11:30 a. m.</b>
</div>
<script>
(function(){{
 const end=Date.now()+{remaining}*1000, out=document.getElementById("countdown");
 function tick(){{
  let n=Math.max(0,Math.floor((end-Date.now())/1000));
  let h=Math.floor(n/3600),m=Math.floor((n%3600)/60),sec=n%60;
  out.textContent=(h?String(h).padStart(2,"0")+" h ":"")+String(m).padStart(2,"0")+" min "+String(sec).padStart(2,"0")+" s";
  if(n<=0) location.reload(); else setTimeout(tick,1000);
 }}
 tick();
}})();
</script>"""
        else:
            countdown_html = ""

        body = f"""
<main class="wrap narrow">
<div class="card">
<div class="actions" style="justify-content:space-between;align-items:center">
<h1 style="margin:0">Menú de {esc(company['name'])}</h1>
{f'<a class="btn secondary" href="/empresa/talma/logout?token={esc(token)}">Cerrar sesión</a>' if is_talma else ''}
</div>
<div class="notice" style="margin:12px 0">
  <b>¿Para qué día quieres hacer tu pedido?</b>
  <form method="get" action="/empresa/{esc(slug)}" style="display:flex;gap:10px;align-items:end;flex-wrap:wrap;margin-top:10px">
    <input type="hidden" name="token" value="{esc(token)}">
    <div>
      <label>Fecha del pedido</label>
      <input type="date" id="fecha-pedido" name="fecha" value="{esc(requested_date)}" min="{today_iso()}" required>
      <div id="dia-pedido" class="muted" style="margin-top:6px;font-weight:700"></div>
    </div>
    <button type="submit" class="btn secondary">Ver menú de ese día</button>
  </form>
  <p class="muted" style="margin-bottom:0">Puedes elegir hoy o una fecha futura. El pedido quedará registrado para la fecha seleccionada.</p>
</div>
<p class="muted">Fecha seleccionada: <b>{esc(fecha_con_dia(requested_date))}</b> · Hora límite: <b>11:30 a. m.</b> (hora de Perú)</p>
{countdown_html}
{notice}
<form method="post" action="/pedido/{esc(slug)}">
<input type="hidden" name="token" value="{esc(token)}">
<input type="hidden" name="fecha" value="{esc(requested_date)}">
{f'<div class="notice ok talma-welcome">Acceso correcto. Hola de nuevo, <b>{esc(talma_employee["nombre"])}</b>.<br><span>Área: {esc(self.display_area("talma", talma_employee["area"]))}</span></div>' if is_talma else ''}
{'' if is_talma else '<div class="grid grid2">\n<div>\n<label>Nombre y apellido</label>\n<input name="nombre" required maxlength="100" placeholder="Ejemplo: Juan Pérez">\n</div>\n<div>\n<label>Área o sede</label>\n<input name="area" required maxlength="80" placeholder="Ejemplo: RAMPA">\n</div>\n</div>'}
{menu_html}
<label>Observación (opcional)</label>
<textarea name="observaciones" maxlength="300" placeholder="Ejemplo: sin cebolla, poco arroz..."></textarea>
{submit}
</form>
</div>
{public_orders_html}

<div class="card no-print" style="margin-top:24px">
<h3>Reporte del día</h3>
<p class="muted">Descarga los pedidos de la fecha seleccionada.</p>
<div class="actions">
<a class="btn secondary" href="/empresa/{esc(slug)}/excel?token={esc(token)}&fecha={esc(requested_date)}">Excel del día</a>
<a class="btn secondary" href="/empresa/{esc(slug)}/pdf?token={esc(token)}&fecha={esc(requested_date)}">PDF del día</a>
</div>
</div>
</main>"""
        self.send_html(page("Realizar pedido", body))

    def submit_order(self, slug: str) -> None:
        form = self.read_form()
        token = form.get("token", "")
        company = get_company(slug, token)
        if not company:
            self.send_html(page("No autorizado", '<main class="wrap narrow"><div class="card"><h1>No autorizado</h1></div></main>'), 403)
            return
        order_date = form.get("fecha", today_iso()).strip()
        is_talma = slug.lower() == "talma"
        talma_employee = None
        if is_talma:
            cookie = self.parse_cookies().get("talma_employee_session")
            talma_employee = valid_talma_employee_session(cookie.value if cookie else None)
            if not talma_employee:
                self.redirect(f"/empresa/talma/login?{urlencode({'token':token})}")
                return
            name = talma_employee["nombre"]
            area = talma_employee["area"]
            dni = talma_employee["dni"]
            login_code = talma_employee["dni"]
            login_email = talma_employee["email"]
        else:
            name = form.get("nombre", "").strip()
            dni = ""
            area = form.get("area", "").strip()
            login_code = ""
            login_email = ""
        entry = form.get("entrada", "").strip()
        main = form.get("fondo", "").strip()
        notes = form.get("observaciones", "").strip()
        delivery = "Entrega en empresa"
        error = None
        try:
            datetime.strptime(order_date, "%Y-%m-%d")
        except ValueError:
            error = "Fecha inválida."
        if order_date < today_iso():
            error = "No se pueden registrar pedidos para una fecha pasada."
        if is_order_closed(order_date):
            error = "Los pedidos del día actual se cierran a las 11:30 a. m. (hora de Perú). Si necesita un pedido fuera de horario, comuníquese por WhatsApp para coordinar la entrega."
        if len(name) < 3:
            error = "Ingrese su nombre y apellido."
        menu = get_menu(order_date)
        valid_entries = {r["name"] for r in menu["entrada"]}
        valid_mains = {r["name"] for r in menu["fondo"]}
        if entry not in valid_entries or main not in valid_mains:
            error = "Seleccione platos válidos del menú publicado."
        if error:
            params = urlencode({"token": token, "fecha": order_date, "error": error})
            self.redirect(f"/empresa/{quote(slug)}?{params}")
            return
        try:
            with db() as conn:
                employee_key = f"talma:{dni}" if is_talma else normalize_key(name)
                conn.execute(
                    """INSERT INTO orders(company_id, order_date, employee_name, employee_key, dni, area, entry_item, main_item, notes, delivery_type, created_at, login_code, login_email)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (company["id"], order_date, name, employee_key, dni, area, entry, main, notes, delivery, now_iso(), login_code, login_email),
                )
        except sqlite3.IntegrityError:
            params = urlencode({"token": token, "fecha": order_date, "error": "Ya existe un pedido para ese usuario en la fecha seleccionada." if is_talma else "Ya existe un pedido con ese nombre en la fecha seleccionada."})
            self.redirect(f"/empresa/{quote(slug)}?{params}")
            return
        self.redirect(f"/empresa/{quote(slug)}?{urlencode({'token':token,'fecha':order_date,'ok':'1'})}")

    def _company_orders(self, company_id: int, order_date: str = ""):
        with db() as conn:
            sql = """SELECT o.id,o.order_date,o.employee_name,o.dni,o.area,o.entry_item,
                            o.main_item,o.notes,o.delivery_type,o.created_at
                     FROM orders o WHERE o.company_id=?"""
            args=[company_id]
            if order_date:
                sql += " AND o.order_date=?"
                args.append(order_date)
            sql += " ORDER BY o.order_date DESC,o.employee_name COLLATE NOCASE,o.id"
            return conn.execute(sql,args).fetchall()

    def company_export_excel(self, slug: str, query: dict[str,str]) -> None:
        token=query.get("token","")
        company=get_company(slug,token)
        if not company:
            self.send_html(page("Enlace inválido",'<main class="wrap narrow"><div class="card"><h1>Enlace inválido</h1></div></main>'),403); return
        requested_date=query.get("fecha","").strip()
        if not requested_date:
            self.send_html(page("Fecha requerida", '<main class="wrap narrow"><div class="card"><h1>Fecha requerida</h1><p>Selecciona una fecha para descargar el reporte.</p></div></main>'), 400)
            return
        rows=self._company_orders(company["id"],requested_date)
        wb=Workbook(); ws=wb.active; ws.title="Pedidos"
        headers=["Fecha","Nombre","DNI","Área / sede","Entrada","Plato de fondo","Observación","Tipo entrega","Hora"]
        ws.append(headers)
        for r in rows:
            ws.append([r["order_date"],r["employee_name"],r["dni"] or "",r["area"] or "",r["entry_item"],r["main_item"],r["notes"] or "",r["delivery_type"] or "",(r["created_at"] or "")[11:16]])
        for i,w in enumerate([14,32,15,20,34,38,45,22,10],1): ws.column_dimensions[get_column_letter(i)].width=w
        for c in ws[1]:
            c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="176B43"); c.alignment=Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for c in row: c.alignment=Alignment(vertical="top",wrap_text=True)
        out=io.BytesIO(); wb.save(out)
        suffix=f"_{requested_date}"
        self.send_bytes(out.getvalue(),"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",f"pedidos_{company['slug']}{suffix}.xlsx")

    def company_export_pdf(self, slug: str, query: dict[str,str]) -> None:
        token=query.get("token","")
        company=get_company(slug,token)
        if not company:
            self.send_html(page("Enlace inválido",'<main class="wrap narrow"><div class="card"><h1>Enlace inválido</h1></div></main>'),403); return
        requested_date=query.get("fecha","").strip()
        if not requested_date:
            self.send_html(page("Fecha requerida", '<main class="wrap narrow"><div class="card"><h1>Fecha requerida</h1><p>Selecciona una fecha para descargar el reporte.</p></div></main>'), 400)
            return
        rows=self._company_orders(company["id"],requested_date)
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer
        out=io.BytesIO()
        doc=SimpleDocTemplate(out,pagesize=landscape(A4),rightMargin=20,leftMargin=20,topMargin=20,bottomMargin=20)
        styles=getSampleStyleSheet()
        title=ParagraphStyle("ExportTitle",parent=styles["Title"],alignment=TA_CENTER,fontSize=18)
        small=ParagraphStyle("Small",parent=styles["BodyText"],fontSize=7,leading=9)
        story=[Paragraph(f"Pedidos — {html.escape(company['name'])}",title),
               Paragraph(f"Fecha: {html.escape(requested_date)}",styles["Normal"]),Spacer(1,8)]
        data=[["Fecha","Nombre","DNI","Área / sede","Entrada","Plato de fondo","Observación","Entrega","Hora"]]
        for r in rows:
            data.append([Paragraph(str(r["order_date"]),small),Paragraph(str(r["employee_name"]),small),Paragraph(str(r["dni"] or ""),small),Paragraph(str(r["area"] or ""),small),Paragraph(str(r["entry_item"]),small),Paragraph(str(r["main_item"]),small),Paragraph(str(r["notes"] or ""),small),Paragraph(str(r["delivery_type"] or ""),small),Paragraph(str((r["created_at"] or "")[11:16]),small)])
        table=Table(data,repeatRows=1,colWidths=[48,95,55,65,100,105,125,65,35])
        table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#176B43")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("GRID",(0,0),(-1,-1),.35,colors.HexColor("#D0D5DD")),("VALIGN",(0,0),(-1,-1),"TOP"),("FONTSIZE",(0,0),(-1,-1),7)]))
        story += [table,Spacer(1,8),Paragraph(f"Total de pedidos: {len(rows)}",styles["Heading3"])]
        doc.build(story); out.seek(0)
        suffix=requested_date or "historial"
        self.send_bytes(out.getvalue(),"application/pdf",f"pedidos_{company['slug']}_{suffix}.pdf")

    def admin_login(self, error: str | None) -> None:
        if self.is_admin():
            self.redirect("/admin/dashboard")
            return
        notice = '<div class="notice error">Contraseña incorrecta.</div>' if error else ""
        body = f"""
<main class="wrap narrow"><div class="card">
<h1>Panel administrador</h1>
{notice}
<form method="post" action="/admin/login">
<label>Contraseña</label><input type="password" name="password" required autofocus>
<button type="submit">Ingresar</button>
</form>
<p class="muted">La contraseña inicial está en el archivo <b>config.json</b>.</p>
</div></main>"""
        self.send_html(page("Administrador", body))

    def login(self) -> None:
        form = self.read_form()
        if hmac.compare_digest(form.get("password", ""), str(CONFIG.get("admin_password", ""))):
            value = session_value()
            secure = "; Secure" if self.headers.get("X-Forwarded-Proto") == "https" else ""
            self.redirect("/admin/dashboard", f"admin_session={value}; Path=/; Max-Age=28800; HttpOnly; SameSite=Lax{secure}")
        else:
            self.redirect("/admin?error=1")

    def admin_personas(self, query: dict[str, str]) -> None:
        company_filter = query.get("empresa", "").strip()
        search = query.get("buscar", "").strip()
        with db() as conn:
            companies = conn.execute("SELECT * FROM companies WHERE active=1 ORDER BY name COLLATE NOCASE").fetchall()
            filters, args = [], []
            if company_filter.isdigit():
                filters.append("o.company_id=?"); args.append(int(company_filter))
            if search:
                term = f"%{search}%"
                filters.append("(o.employee_name LIKE ? OR o.dni LIKE ? OR o.area LIKE ?)")
                args.extend([term, term, term])
            where = (" WHERE " + " AND ".join(filters)) if filters else ""
            rows = conn.execute(
                f"""SELECT c.name AS company_name, o.employee_name, o.dni, MAX(o.area) AS area,
                           COUNT(*) AS total_consumo, MIN(o.order_date) AS primer_pedido,
                           MAX(o.order_date) AS ultimo_pedido
                    FROM orders o JOIN companies c ON c.id=o.company_id
                    {where}
                    GROUP BY c.id, o.employee_key, o.employee_name, o.dni
                    ORDER BY c.name COLLATE NOCASE, o.employee_name COLLATE NOCASE""", args).fetchall()

        options = '<option value="">Todas las empresas</option>' + "".join(
            f'<option value="{c["id"]}" {"selected" if str(c["id"])==company_filter else ""}>{esc(c["name"])}</option>' for c in companies
        )
        rows_html = "".join(
            f'<tr><td><b>{esc(r["employee_name"])}</b></td><td>{esc(r["company_name"])}</td>'
            f'<td>{esc(r["dni"]) if r["dni"] else "—"}</td><td>{esc(r["area"]) if r["area"] else "—"}</td>'
            f'<td><strong>{r["total_consumo"]}</strong></td><td>{esc(r["primer_pedido"])}</td><td>{esc(r["ultimo_pedido"])}</td></tr>'
            for r in rows
        ) or '<tr><td colspan="7">No se encontraron personas.</td></tr>'

        body=f"""
<main class="wrap">
<div class="actions no-print" style="justify-content:space-between;margin-bottom:16px"><h1 style="margin:0"> Base de datos · Personas y consumo</h1><a href="/admin/dashboard">← Pedidos</a></div>
<div class="card no-print"><h2>Buscar persona</h2>
<form method="get" action="/admin/personas" class="grid grid3">
<div><label>Nombre, DNI o área</label><input name="buscar" value="{esc(search)}" placeholder="Ej. Juan, 76543210, RAMPA"></div>
<div><label>Empresa</label><select name="empresa">{options}</select></div>
<div style="align-self:end"><button>Buscar</button> <a class="btn secondary" href="/admin/personas">Limpiar</a></div>
</form></div>
<div class="grid grid3"><div class="stat">Personas encontradas<b>{len(rows)}</b></div>
<div class="stat">Empresas<b>{len(companies)}</b></div><div class="stat">Consumo total<b>{sum(int(r["total_consumo"]) for r in rows)}</b></div></div>
<div class="card"><h2>Consumo por persona</h2>
<p class="muted">TALMA se identifica principalmente por DNI; las demás empresas por nombre.</p>
<div class="table-wrap"><table><thead><tr><th>Nombre</th><th>Empresa</th><th>DNI</th><th>Área</th><th>Total almuerzos</th><th>Primer pedido</th><th>Último pedido</th></tr></thead>
<tbody>{rows_html}</tbody></table></div></div></main>"""
        self.send_html(page("Base de datos · Personas", body))

    def admin_resumen(self, query: dict[str, str]) -> None:
        selected_date = query.get("fecha", today_iso())
        try:
            datetime.strptime(selected_date, "%Y-%m-%d")
        except ValueError:
            selected_date = today_iso()
        with db() as conn:
            company_rows = conn.execute(
                """SELECT c.id, c.name, c.slug, COUNT(o.id) AS total
                   FROM companies c
                   LEFT JOIN orders o ON o.company_id=c.id AND o.order_date=?
                   WHERE c.active=1 GROUP BY c.id
                   ORDER BY CASE c.slug WHEN 'liderman' THEN 1 WHEN 'aeropuerto' THEN 2
                   WHEN 'talma' THEN 3 WHEN 'policia' THEN 4 ELSE 5 END, c.name""",
                (selected_date,)).fetchall()
        grand_total=sum(int(r["total"]) for r in company_rows)
        cards="".join(f'<div class="big-company-card"><div class="company-name">{esc(r["name"])}</div><div class="company-total">{r["total"]}</div><div class="company-label">MENÚS / PEDIDOS</div></div>' for r in company_rows)
        body=f"""
<main class="wrap">
<div class="actions no-print" style="justify-content:space-between;margin-bottom:16px"><h1 style="margin:0"> Resumen de almuerzos</h1><a href="/admin/dashboard">← Panel</a></div>
<div class="card no-print"><form method="get" action="/admin/resumen" class="actions" style="align-items:end">
<div><label>Fecha</label><input type="date" name="fecha" value="{esc(selected_date)}" required></div><button>Actualizar</button>
<a class="btn secondary" href="/admin/resumen?fecha={today_iso()}">Hoy</a></form></div>
<div class="hero-total"><div class="hero-kicker">TOTAL GENERAL</div><div class="hero-number">{grand_total}</div><div class="hero-date">{esc(fecha_con_dia(selected_date))}</div></div>
<div class="company-cards">{cards}</div>
<div class="card" style="text-align:center"><h2>Resumen de {esc(fecha_con_dia(selected_date))}</h2><p class="muted">Esta pantalla se actualiza automáticamente cada 30 segundos.</p></div>
</main><script>setTimeout(function(){{ location.reload(); }},30000);</script>"""
        extra="""<style>
.hero-total{background:linear-gradient(135deg,#101828,#175cd3);color:white;border-radius:24px;padding:32px;text-align:center;margin-bottom:24px;box-shadow:0 14px 40px rgba(16,24,40,.22)}
.hero-kicker{font-weight:800;letter-spacing:3px;font-size:14px;opacity:.9}.hero-number{font-size:72px;font-weight:900;line-height:1;margin:10px 0}.hero-date{font-size:20px;font-weight:700}
.company-cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:20px;margin-bottom:24px}.big-company-card{background:white;border:1px solid #e4e7ec;border-radius:22px;padding:26px;text-align:center;box-shadow:0 8px 24px rgba(16,24,40,.08)}.company-name{font-size:25px;font-weight:900}.company-total{font-size:64px;font-weight:900;margin:12px 0;color:#175cd3}.company-label{font-size:13px;font-weight:800;letter-spacing:1.5px;color:#667085}
</style>"""
        self.send_html(page("Resumen de almuerzos", body, extra))

    def admin_dashboard(self, query: dict[str, str]) -> None:
        selected_date = query.get("fecha", today_iso())
        try:
            datetime.strptime(selected_date, "%Y-%m-%d")
        except ValueError:
            selected_date = today_iso()
        selected_company = query.get("empresa", "")
        with db() as conn:
            companies = conn.execute(
                """SELECT * FROM companies
                   ORDER BY CASE slug
                       WHEN 'liderman' THEN 1
                       WHEN 'aeropuerto' THEN 2
                       WHEN 'talma' THEN 3
                       WHEN 'policia' THEN 4
                       ELSE 5
                   END, name"""
            ).fetchall()
            filters = ["o.order_date=?"]
            args: list[object] = [selected_date]
            if selected_company.isdigit():
                filters.append("o.company_id=?")
                args.append(int(selected_company))
            orders = conn.execute(
                f"""SELECT o.*, c.name AS company_name FROM orders o
                    JOIN companies c ON c.id=o.company_id
                    WHERE {' AND '.join(filters)} ORDER BY o.employee_name COLLATE NOCASE, c.name COLLATE NOCASE, o.created_at""",
                args,
            ).fetchall()
            total_today = conn.execute("SELECT COUNT(*) FROM orders WHERE order_date=?", (selected_date,)).fetchone()[0]
            company_total = conn.execute("SELECT COUNT(*) FROM companies WHERE active=1").fetchone()[0]
            entry_summary = conn.execute(
                "SELECT entry_item, COUNT(*) qty FROM orders WHERE order_date=? GROUP BY entry_item ORDER BY qty DESC, entry_item",
                (selected_date,),
            ).fetchall()
            main_summary = conn.execute(
                "SELECT main_item, COUNT(*) qty FROM orders WHERE order_date=? GROUP BY main_item ORDER BY qty DESC, main_item",
                (selected_date,),
            ).fetchall()
        menu = get_menu(selected_date)
        entries_text = "\n".join(r["name"] for r in menu["entrada"])
        mains_text = "\n".join(r["name"] for r in menu["fondo"])
        entry_menu_view = "".join(f"<li>{esc(r['name'])}</li>" for r in menu["entrada"]) or "<li><span class='muted'>No hay entradas publicadas para este día.</span></li>"
        main_menu_view = "".join(f"<li>{esc(r['name'])}</li>" for r in menu["fondo"]) or "<li><span class='muted'>No hay platos de fondo publicados para este día.</span></li>" 
        host = self.headers.get("Host", f"localhost:{PORT}")
        scheme = "https" if self.headers.get("X-Forwarded-Proto") == "https" else "http"

        company_options = '<option value="">Todas</option>' + "".join(
            f'<option value="{c["id"]}" {"selected" if str(c["id"])==selected_company else ""}>{esc(c["name"])}</option>' for c in companies
        )
        company_rows = "".join(
            f"""<tr><td>{esc(c['name'])}</td><td><code>{esc(scheme)}://{esc(host)}/empresa/{esc(c['slug'])}?token={esc(c['token'])}</code></td>
<td>{'Activa' if c['active'] else 'Inactiva'}</td><td><form method="post" action="/admin/empresa/toggle"><input type="hidden" name="id" value="{c['id']}"><button class="btn small secondary">{'Desactivar' if c['active'] else 'Activar'}</button></form></td></tr>"""
            for c in companies
        ) or '<tr><td colspan="4">No hay empresas.</td></tr>'
        order_rows = "".join(
            f"""<tr><td>{esc(o['company_name'])}</td><td>{esc(o['dni']) if o['company_name'].lower() == 'talma' and o['dni'] else '—'}</td><td>{esc(o['employee_name'])}</td><td>{esc(o['area'])}</td><td>{esc(o['entry_item'])}</td><td>{esc(o['main_item'])}</td><td>{esc(o['delivery_type'])}</td><td>{esc(o['notes'])}</td><td>{esc(o['created_at'][11:16])}</td>
<td><form method="post" action="/admin/pedido/eliminar" onsubmit="return confirm('¿Eliminar pedido?')"><input type="hidden" name="id" value="{o['id']}"><input type="hidden" name="fecha" value="{esc(selected_date)}"><button class="btn small danger">Eliminar</button></form></td></tr>"""
            for o in orders
        ) or '<tr><td colspan="10">No hay pedidos para el filtro seleccionado.</td></tr>'
        entry_rows = "".join(f"<li>{esc(r['entry_item'])}: <b>{r['qty']}</b></li>" for r in entry_summary) or "<li>Sin pedidos</li>"
        main_rows = "".join(f"<li>{esc(r['main_item'])}: <b>{r['qty']}</b></li>" for r in main_summary) or "<li>Sin pedidos</li>"
        export_params = urlencode({"fecha": selected_date, "empresa": selected_company})
        if query.get("ok"):
            notice = f'<div class="notice ok"> Menú guardado correctamente para <b>{esc(selected_date)}</b>. Se volvió a leer desde la base de datos para comprobar que quedó almacenado.</div>'
        elif query.get("menu_error"):
            notice = f'<div class="notice error">{esc(query["menu_error"])}</div>'
        else:
            notice = ""

        body = f"""
<main class="wrap">
<div class="actions no-print" style="justify-content:space-between;margin-bottom:16px"><h1 style="margin:0">Panel administrador</h1><a href="/admin/logout">Cerrar sesión</a></div>
<div class="admin-tabs no-print"><a class="active" href="/admin/dashboard">Pedidos</a><a href="/admin/personas"> Personas y consumo</a><a href="/admin/resumen"> Resumen visual</a><a href="/admin/talma/">TALMA</a><a href="/admin/policia/">POLICÍA</a></div>
{notice}
<div class="grid grid3">
<div class="stat">Pedidos del día<b>{total_today}</b></div><div class="stat">Empresas activas<b>{company_total}</b></div><div class="stat">Fecha seleccionada<b style="font-size:18px">{esc(selected_date)}</b></div>
</div>

<div class="card no-print">
<h2> Consultar menú de cualquier día</h2>
<form method="get" action="/admin/dashboard" class="actions" style="align-items:end;gap:12px;flex-wrap:wrap">
  <div>
    <label>Selecciona una fecha</label>
    <input type="date" name="fecha" value="{esc(selected_date)}" required>
  </div>
  <input type="hidden" name="empresa" value="{esc(selected_company)}">
  <button type="submit">Ver menú</button>
</form>
<div class="grid grid2" style="margin-top:14px">
  <div>
    <h3>Entradas — {esc(fecha_con_dia(selected_date))}</h3>
    <ul>{entry_menu_view}</ul>
  </div>
  <div>
    <h3>Platos de fondo — {esc(fecha_con_dia(selected_date))}</h3>
    <ul>{main_menu_view}</ul>
  </div>
</div>
<p class="muted" style="margin-bottom:0">Al cambiar la fecha y pulsar <b>Ver menú</b>, se carga directamente el menú guardado para ese día desde la base de datos.</p>
</div>

<div class="card no-print">
<h2>Menú diario</h2>
<form method="post" action="/admin/menu">
<div class="grid grid2"><div><label>Fecha</label><input type="date" name="fecha" value="{esc(selected_date)}" required></div><div><p class="muted">Escriba un plato por línea. Al guardar, reemplaza el menú de esa fecha.</p></div></div>
<div class="grid grid2"><div><label>Entradas</label><textarea name="entradas" required>{esc(entries_text)}</textarea></div><div><label>Platos de fondo</label><textarea name="fondos" required>{esc(mains_text)}</textarea></div></div>
<button>Guardar menú de esta fecha</button>
</form>
<div class="notice" style="margin-top:10px">
  <b>Menú cargado para {esc(selected_date)}</b><br>
  Entradas guardadas: <b>{len(menu["entrada"])}</b> · Platos de fondo guardados: <b>{len(menu["fondo"])}</b>.
  <a href="/admin/dashboard?fecha={esc(selected_date)}">Volver a cargar esta fecha</a>
</div>
</div>

<div class="card no-print">
<h2>Empresas y enlaces privados</h2>
<form method="post" action="/admin/empresa" class="actions"><input name="nombre" required placeholder="Nombre de la nueva empresa" style="max-width:420px"><button>Crear empresa</button></form>
<div class="table-wrap"><table><thead><tr><th>Empresa</th><th>Enlace para empleados</th><th>Estado</th><th></th></tr></thead><tbody>{company_rows}</tbody></table></div>
</div>



<div class="card">
<h2>Pedidos y resumen</h2>
<form method="get" action="/admin/dashboard" class="grid grid3 no-print">
<div><label>Fecha</label><input type="date" name="fecha" value="{esc(selected_date)}"></div><div><label>Empresa</label><select name="empresa">{company_options}</select></div><div style="align-self:end"><button>Aplicar filtro</button></div>
</form>
<div class="grid grid2"><div><h3>Entradas</h3><ul>{entry_rows}</ul></div><div><h3>Fondos</h3><ul>{main_rows}</ul></div></div>
<div class="actions no-print" style="margin-bottom:14px"><a class="btn secondary" href="/admin/export.csv?{export_params}">Descargar CSV/Excel</a><a class="btn" target="_blank" href="/admin/cupones?{export_params}">Generar cupones</a></div>
<div class="table-wrap"><table><thead><tr><th>Empresa</th><th>DNI</th><th>Empleado</th><th>Área</th><th>Entrada</th><th>Fondo</th><th>Modalidad</th><th>Observación</th><th>Hora</th><th></th></tr></thead><tbody>{order_rows}</tbody></table></div>
</div>

<div class="card no-print" id="excel-historico">
<h2> Actualizar un Excel pasado / reporte histórico</h2>
<p class="muted">Esta opción está dentro del panel de <b>Pedidos</b>. Sube aquí un Excel de días o meses anteriores y el sistema lo combinará con los pedidos que ya están registrados actualmente.</p>
<div class="grid grid2">
  <div>
    <h3>1. Seleccionar Excel antiguo</h3>
    <form method="post" action="/admin/historico-excel" enctype="multipart/form-data">
      <input type="file" name="excel_historico" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" required style="width:100%;box-sizing:border-box;margin-bottom:10px">
      <button type="submit">Actualizar Excel y generar reporte</button>
    </form>
  </div>
  <div>
    <h3>2. ¿Qué genera?</h3>
    <ul>
      <li><b>DETALLE UNIFICADO:</b> todos los pedidos antiguos + actuales.</li>
      <li><b>Orden alfabético:</b> las personas aparecen de A a Z.</li>
      <li><b>N° Almuerzo:</b> se acumula por persona.</li>
      <li><b>REPORTE PEDIDOS:</b> totales por persona, día y plato.</li>
    </ul>
  </div>
</div>
<p class="muted" style="margin-bottom:0"><b>Importante:</b> el archivo original no se modifica. Se descarga una copia nueva con el reporte actualizado.</p>
</div>
</main>"""
        self.send_html(page("Panel administrador", body))

    def save_menu(self) -> None:
        """Guarda de forma atómica el menú de una fecha y verifica inmediatamente lo almacenado."""
        form = self.read_form()
        raw_date = form.get("fecha", "").strip()

        # Acepta tanto el valor nativo del <input type="date"> como DD/MM/YYYY
        # por si el navegador/formulario envía una fecha localizada.
        menu_date = raw_date
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                parsed_date = datetime.strptime(raw_date, fmt)
                menu_date = parsed_date.strftime("%Y-%m-%d")
                break
            except ValueError:
                continue
        else:
            params = urlencode({
                "fecha": today_iso(),
                "menu_error": "Fecha inválida. Seleccione una fecha válida del calendario."
            })
            self.redirect(f"/admin/dashboard?{params}")
            return

        def unique_dishes(raw_text: str) -> list[str]:
            result: list[str] = []
            seen: set[str] = set()
            for line in (raw_text or "").splitlines():
                # También acepta platos separados por ';' o por salto de línea.
                for part in re.split(r"[;\n]+", line):
                    dish = re.sub(r"\s+", " ", part).strip()
                    if not dish:
                        continue
                    dish = dish[:120]
                    key = normalize_key(dish)
                    if key and key not in seen:
                        seen.add(key)
                        result.append(dish)
            return result

        entries = unique_dishes(form.get("entradas", ""))
        mains = unique_dishes(form.get("fondos", ""))

        if not entries or not mains:
            params = urlencode({
                "fecha": menu_date,
                "menu_error": "Debes ingresar al menos una entrada y un plato de fondo para esa fecha."
            })
            self.redirect(f"/admin/dashboard?{params}")
            return

        try:
            with db() as conn:
                # Una única transacción: nunca queda una fecha parcialmente guardada.
                conn.execute("BEGIN IMMEDIATE")
                conn.execute("DELETE FROM menu_items WHERE menu_date=?", (menu_date,))
                conn.executemany(
                    "INSERT INTO menu_items(menu_date, category, name) VALUES (?, 'entrada', ?)",
                    [(menu_date, dish) for dish in entries],
                )
                conn.executemany(
                    "INSERT INTO menu_items(menu_date, category, name) VALUES (?, 'fondo', ?)",
                    [(menu_date, dish) for dish in mains],
                )

                # Verificación antes del COMMIT.
                stored_entries = conn.execute(
                    "SELECT name FROM menu_items WHERE menu_date=? AND category='entrada' AND active=1 ORDER BY id",
                    (menu_date,),
                ).fetchall()
                stored_mains = conn.execute(
                    "SELECT name FROM menu_items WHERE menu_date=? AND category='fondo' AND active=1 ORDER BY id",
                    (menu_date,),
                ).fetchall()

                saved_entries = [r["name"] for r in stored_entries]
                saved_mains = [r["name"] for r in stored_mains]
                if saved_entries != entries or saved_mains != mains:
                    raise sqlite3.Error("La verificación del menú guardado no coincide con los datos enviados.")

                conn.commit()

        except sqlite3.Error as error:
            print(f"Error al guardar/verificar el menú de {menu_date}: {error}", file=sys.stderr)
            params = urlencode({
                "fecha": menu_date,
                "menu_error": f"No se pudo guardar el menú del {menu_date}. Inténtelo nuevamente."
            })
            self.redirect(f"/admin/dashboard?{params}")
            return

        # Redirige a la misma fecha y muestra el menú recién guardado.
        self.redirect(
            f"/admin/dashboard?{urlencode({'fecha': menu_date, 'ok': '1', 'menu_guardado': '1'})}"
        )

    def create_company(self) -> None:
        form = self.read_form()
        name = form.get("nombre", "").strip()
        if not name:
            self.redirect("/admin/dashboard")
            return
        base = slugify(name)
        slug = base
        with db() as conn:
            n = 2
            while conn.execute("SELECT 1 FROM companies WHERE slug=?", (slug,)).fetchone():
                slug = f"{base}-{n}"
                n += 1
            conn.execute(
                "INSERT INTO companies(name, slug, token, created_at) VALUES (?, ?, ?, ?)",
                (name[:120], slug, secrets.token_urlsafe(18), now_iso()),
            )
        self.redirect("/admin/dashboard?ok=1")

    def toggle_company(self) -> None:
        form = self.read_form()
        if form.get("id", "").isdigit():
            with db() as conn:
                conn.execute("UPDATE companies SET active=CASE active WHEN 1 THEN 0 ELSE 1 END WHERE id=?", (int(form["id"]),))
        self.redirect("/admin/dashboard?ok=1")

    def update_historical_excel(self) -> None:
        try:
            filename, data = read_multipart_file(self, "excel_historico")
            if not filename.lower().endswith(".xlsx"):
                raise ValueError("Solo se aceptan archivos .xlsx.")
            workbook = load_workbook(io.BytesIO(data))
            historical = extract_historical_orders(workbook)

            with db() as conn:
                current_rows = conn.execute(
                    """SELECT o.order_date, c.name AS company_name, o.employee_name, o.area,
                              o.entry_item, o.main_item, o.notes
                       FROM orders o JOIN companies c ON c.id=o.company_id
                       ORDER BY o.order_date, c.name, o.employee_name"""
                ).fetchall()

            current = [{
                "fecha": r["order_date"],
                "empresa": r["company_name"],
                "nombre": r["employee_name"],
                "dni": r["dni"] if "dni" in r.keys() else "",
                "area": r["area"],
                "entrada": r["entry_item"],
                "fondo": r["main_item"],
                "observacion": r["notes"],
            } for r in current_rows]

            total, people = build_historical_report(workbook, historical, current)
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(filename).stem) or "pedidos_historico"
            download_name = f"{safe_name}_reporte_actualizado.xlsx"
            self.send_bytes(output.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            download_name)
        except Exception as error:
            print(f"Error al actualizar Excel histórico: {error}", file=sys.stderr)
            message = esc(str(error))
            self.send_html(page("Error de Excel", f'<main class="wrap narrow"><div class="card"><h1>No se pudo actualizar el Excel</h1><p>{message}</p><a class="btn" href="/admin/dashboard">Volver al panel</a></div></main>'), 400)

    def delete_order(self) -> None:
        form = self.read_form()
        if form.get("id", "").isdigit():
            with db() as conn:
                conn.execute("DELETE FROM orders WHERE id=?", (int(form["id"]),))
        self.redirect(f"/admin/dashboard?{urlencode({'fecha':form.get('fecha',today_iso()),'ok':'1'})}")

    def filtered_orders(self, query: dict[str, str]):
        selected_date = query.get("fecha", today_iso())
        company = query.get("empresa", "")
        filters = ["o.order_date=?"]
        args: list[object] = [selected_date]
        if company.isdigit():
            filters.append("o.company_id=?")
            args.append(int(company))
        with db() as conn:
            rows = conn.execute(
                f"""SELECT o.*, c.name company_name FROM orders o JOIN companies c ON c.id=o.company_id
                    WHERE {' AND '.join(filters)} ORDER BY o.employee_name COLLATE NOCASE, c.name COLLATE NOCASE, o.created_at""",
                args,
            ).fetchall()
        return selected_date, rows

    def export_csv(self, query: dict[str, str]) -> None:
        selected_date, rows = self.filtered_orders(query)
        output = io.StringIO(newline="")
        writer = csv.writer(output, delimiter=";")
        writer.writerow(["Empresa", "DNI", "Empleado", "Área", "Entrada", "Plato de fondo", "Modalidad", "Observación", "Fecha", "Hora"])
        for r in rows:
            dni = r["dni"] if r["company_name"].lower() == "talma" else ""
            writer.writerow([r["company_name"], dni, r["employee_name"], r["area"], r["entry_item"], r["main_item"], r["delivery_type"], r["notes"], r["order_date"], r["created_at"][11:16]])
        data = ("\ufeff" + output.getvalue()).encode("utf-8")
        self.send_bytes(data, "text/csv; charset=utf-8", f"pedidos-{selected_date}.csv")

    def coupons(self, query: dict[str, str]) -> None:
        selected_date, rows = self.filtered_orders(query)

        def dish_style(value: str, kind: str = "entry") -> str:
            # Ajusta automáticamente el tamaño para que ambos platos se vean grandes,
            # centrados y sin cortarse, respetando tildes y textos largos.
            length = len(value.strip())
            if kind == "entry":
                if length > 34:
                    return "font-size:22px;line-height:1.04"
                if length > 24:
                    return "font-size:27px;line-height:1.04"
                return "font-size:33px;line-height:1.02"
            if length > 42:
                return "font-size:18px;line-height:1.06"
            if length > 30:
                return "font-size:21px;line-height:1.06"
            return "font-size:25px;line-height:1.04"

        ticket_cards = []
        for idx, r in enumerate(rows, start=1):
            ticket_cards.append(
                f"""<section class="ticket">
<div class="ticket-top"><div class="ticket-number">TICKET {idx:03d}</div><div class="ticket-company">{esc((r['company_name'] or '').upper())}</div></div>
<div class="ticket-head">
  <div class="ticket-area">{esc(self.display_area((r['company_name'] or ''), (r['area'] or 'Sin área')).upper())}</div>
  <div class="ticket-person">{esc((r['employee_name'] or '').upper())}</div>
</div>
<div class="dish-entry" style="{dish_style(r['entry_item'], 'entry')}">{esc(r['entry_item'])}</div>
<div class="dish-main" style="{dish_style(r['main_item'], 'main')}">{esc(r['main_item'])}</div>
<div class="ticket-notes-title">OBSERVACIÓN</div>
<div class="ticket-notes">{esc(r['notes'] or '—')}</div>
</section>"""
            )

        pages = []
        for i in range(0, len(ticket_cards), 8):
            pages.append(f'<div class="ticket-page">{"".join(ticket_cards[i:i+8])}</div>')
        tickets = "".join(pages) or '<div class="notice error">No hay pedidos para imprimir.</div>'

        coupon_css = """
<style>
.coupon-wrap{max-width:1180px;margin:18px auto;padding:0 14px}
.coupon-card{background:#fff;border:1px solid var(--line);border-radius:14px;padding:16px;box-shadow:0 2px 10px rgba(16,24,40,.04)}
.ticket-page{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));grid-template-rows:repeat(4,1fr);gap:0;margin:12px auto 24px;max-width:190mm;min-height:273mm;background:#fff;border:1px solid #222}
.ticket{margin:0!important;border:1px solid #222!important;padding:4.5mm 4.5mm 4mm!important;min-width:0;min-height:0;display:flex;flex-direction:column;justify-content:flex-start;text-align:left;overflow:hidden;background:#fff}
.ticket-top{display:flex;justify-content:space-between;align-items:flex-start;min-height:16px}
.ticket-number{font-size:9px;font-weight:900;letter-spacing:.3px}
.ticket-company{font-size:8px;line-height:1;font-weight:900;letter-spacing:.5px;border:1px solid #222;padding:1px 3px;white-space:nowrap}
.ticket-head{text-align:center;margin-top:0;margin-bottom:8px}
.ticket-area{font-size:14px;line-height:1.02;font-weight:800;white-space:normal;overflow-wrap:anywhere}
.ticket-person{font-size:14px;line-height:1.02;font-weight:800;white-space:normal;overflow-wrap:anywhere;margin-top:1px}
.dish-entry,.dish-main{font-weight:1000;text-align:center;overflow-wrap:anywhere;word-break:normal;margin:0 auto}
.dish-entry{margin-top:4px;margin-bottom:22px}
.dish-main{margin-bottom:22px}
.ticket-notes-title{font-size:10px;line-height:1.05;font-weight:1000;margin-top:auto;margin-bottom:2px}
.ticket-notes{font-size:10px;line-height:1.15;font-weight:700;white-space:normal;overflow-wrap:anywhere}
@media(max-width:760px){.ticket-page{grid-template-columns:1fr;grid-template-rows:none;min-height:0}.ticket{min-height:250px}}
@page{size:A4 portrait;margin:8mm}
@media print{
  html,body{margin:0!important;padding:0!important;background:#fff!important}
  .coupon-wrap{max-width:none!important;margin:0!important;padding:0!important}
  .coupon-card{border:0!important;border-radius:0!important;box-shadow:none!important;padding:0!important;margin:0!important}
  .ticket-page{
    display:grid!important;
    width:194mm!important;
    max-width:194mm!important;
    height:281mm!important;
    min-height:281mm!important;
    box-sizing:border-box!important;
    grid-template-columns:repeat(2,minmax(0,1fr))!important;
    grid-template-rows:repeat(4,minmax(0,1fr))!important;
    grid-auto-rows:calc(281mm / 4)!important;
    gap:0!important;
    margin:0!important;
    page-break-after:always;
    break-after:page;
  }
  .ticket-page:last-child{page-break-after:auto;break-after:auto}
  .ticket{
    width:auto!important;
    height:auto!important;
    min-height:0!important;
    max-height:none!important;
    box-sizing:border-box!important;
    break-inside:avoid;
    page-break-inside:avoid;
  }
}
</style>
"""
        body = f"""
<main class="coupon-wrap"><div class="coupon-card">
<div class="actions no-print" style="margin-bottom:10px"><button onclick="window.print()">Imprimir / Guardar PDF</button><a class="btn secondary" href="/admin/dashboard?fecha={esc(selected_date)}">Volver</a></div>
<h1 class="no-print">Cupones — 8 por hoja</h1>
<p class="muted no-print" style="margin-top:-6px">Diseño ajustado al modelo solicitado, sin fecha y sin las etiquetas “Entrada” ni “Plato de fondo”.</p>
{tickets}
</div></main>"""
        self.send_html(page("Cupones", body, coupon_css))


def run() -> None:
    global PORT
    if len(sys.argv) > 1:
        try:
            PORT = int(sys.argv[1])
        except ValueError:
            print("Puerto inválido; usando 8080.")
            PORT = 8080
    init_db()
    server = ThreadingHTTPServer((HOST, PORT), AppHandler)
    print("=" * 62)
    print(f"Sistema iniciado: http://localhost:{PORT}")
    print(f"Panel administrador: http://localhost:{PORT}/admin")
    print(f"Contraseña inicial: {CONFIG.get('admin_password')}")
    print("Presione Ctrl+C para detener.")
    print("=" * 62)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServidor detenido.")
    finally:
        server.server_close()


if __name__ == "__main__":
    run()


WEEKDAYS_ES = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]

def fecha_con_dia(fecha_iso):
    try:
        d = datetime.strptime(str(fecha_iso), "%Y-%m-%d").date()
        meses = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"]
        return f"{WEEKDAYS_ES[d.weekday()]} {d.day} de {meses[d.month-1]} de {d.year}"
    except Exception:
        return str(fecha_iso)

